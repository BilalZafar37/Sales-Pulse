# cust_profiles_bp.py
from __future__ import annotations
import os, io, json, hashlib, tempfile
from datetime import datetime, timedelta
from typing import Dict, Any, List, Tuple, Optional

from flask import (
    Blueprint, request, render_template, redirect, url_for, flash, send_file,
    jsonify
)
from werkzeug.utils import secure_filename

from sqlalchemy import text, and_, or_, func
from sqlalchemy.exc import SQLAlchemyError

from models import (
    model, Base,
    SP_Customer, SP_SKU, SP_Customer_SKU_Map, SP_CategoriesMappingMain,
    SP_CustomerUploadProfile, SP_CustomerUploadProfileDetail, SP_SellOut_Staging,
    SP_SellOutUploads
)
from config import STATIC_DIR, BASE_DIR

# Optional: if you have login/roles
# from flask_login import login_required, current_user

bp = Blueprint("sellout_profiles", __name__, static_folder=STATIC_DIR, url_prefix="/sellout/profiles")

# ---------------------------
# Helpers (file, detect, parse)
# ---------------------------

ALLOWED_EXTS = {".xlsx", ".xlsm", ".xltx", ".xltm"}

# For GUI selection
import uuid
TMP_UPLOAD_DIR = os.path.join(BASE_DIR, "tmp_uploads")
os.makedirs(TMP_UPLOAD_DIR, exist_ok=True)

def _save_temp_upload(file_storage) -> tuple[str, str]:
    token = uuid.uuid4().hex
    fname = secure_filename(file_storage.filename)
    fpath = os.path.join(TMP_UPLOAD_DIR, f"{token}__{fname}")
    file_storage.save(fpath)
    return token, fpath

def _resolve_token_path(token: str) -> str | None:
    # naive: find file starting with token__
    for nm in os.listdir(TMP_UPLOAD_DIR):
        if nm.startswith(token + "__"):
            return os.path.join(TMP_UPLOAD_DIR, nm)
    return None

def _sheet_headers(fpath: str, sheet: str, header_row: int) -> list[dict]:
    """
    Return [{'col':1,'excel':'A','header':'Qty'}, ...] using the given header_row.
    If no header text, just show A,B,C... without header.
    """
    from openpyxl.utils import get_column_letter
    wb = _load_workbook(fpath)
    ws = wb[sheet]
    max_cols = min(ws.max_column or 30, 60)
    cols = []
    for c in range(1, max_cols + 1):
        v = ws.cell(row=header_row, column=c).value if header_row else None
        label = str(v).strip() if isinstance(v, str) and v else (v if v is not None else "")
        cols.append({"col": c, "excel": get_column_letter(c), "header": label})
    return cols


def _allowed_file(filename: str) -> bool:
    ext = os.path.splitext(filename)[1].lower()
    return ext in ALLOWED_EXTS

def _file_hash(fpath: str) -> str:
    h = hashlib.sha256()
    with open(fpath, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()

def _load_workbook(fpath: str):
    # Lazy import openpyxl to avoid global dependency if not used
    from openpyxl import load_workbook
    wb = load_workbook(fpath, data_only=True)
    return wb

def _guess_sheet_and_header(fpath: str) -> Tuple[str, int, int]:
    """
    Returns (sheet_name, header_row_index, data_start_row)
    Simple heuristics: pick the sheet with the largest 'dense' block;
    header row is first row with many strings; data starts next row.
    """
    wb = _load_workbook(fpath)
    best_sheet = wb.sheetnames[0]
    best_score = -1
    best_header = 1
    best_start = 2

    for s in wb.sheetnames:
        ws = wb[s]
        max_rows = min(ws.max_row or 200, 200)   # cap scan
        max_cols = min(ws.max_column or 30, 30)

        # density score: count non-empty cells in first 50 rows
        score = 0
        header_row = 1
        for r in range(1, min(max_rows, 50) + 1):
            row_vals = []
            non_empty = 0
            text_like = 0
            for c in range(1, max_cols + 1):
                v = ws.cell(row=r, column=c).value
                if v not in (None, ""):
                    non_empty += 1
                    if isinstance(v, str):
                        row_vals.append(v.strip())
                        if v.strip():
                            text_like += 1
            # naive header heuristic: more strings than numbers and at least 3 distinct tokens
            if header_row == 1 and text_like >= 2 and non_empty >= 2:
                header_row = r
                # do not break; still compute density
            score += non_empty

        data_start = header_row + 1 if header_row else 2

        if score > best_score:
            best_sheet, best_score = s, score
            best_header, best_start = header_row, data_start

    return best_sheet, best_header, best_start

def _suggest_columns(fpath: str, sheet: str, header_row: int, data_start: int) -> Dict[str, Any]:
    """
    Suggest likely columns for date/article/qty/customer_sku by sampling first ~50 rows.
    """
    from dateutil.parser import parse as date_parse
    wb = _load_workbook(fpath)
    ws = wb[sheet]
    max_rows = min(ws.max_row or 200, 200)
    max_cols = min(ws.max_column or 30, 30)

    candidates = {"date": None, "article": None, "qty": None, "customer_sku": None}
    date_hits = [0] * (max_cols + 1)
    qty_hits  = [0] * (max_cols + 1)
    sku_len   = [0] * (max_cols + 1)

    for r in range(data_start, min(max_rows, data_start + 50) + 1):
        for c in range(1, max_cols + 1):
            v = ws.cell(row=r, column=c).value
            if v in (None, ""):
                continue
            # date-ish
            if isinstance(v, (int, float)) and v > 59:  # excel serial (rough)
                date_hits[c] += 1
            else:
                try:
                    date_parse(str(v)); date_hits[c] += 1
                except Exception:
                    pass
            # qty-ish
            try:
                q = float(str(v).replace(",", "").replace(" ", "").replace("(", "-").replace(")", ""))
                # consider positive dominance
                if q != 0:
                    qty_hits[c] += 1
            except Exception:
                pass
            # sku-ish (alphanumeric strings 4–20 chars)
            if isinstance(v, str):
                s = v.strip()
                if 4 <= len(s) <= 20:
                    sku_len[c] += 1

    # choose max indices
    if any(date_hits): candidates["date"] = int(max(range(1, max_cols + 1), key=lambda i: date_hits[i]))
    if any(qty_hits):  candidates["qty"]  = int(max(range(1, max_cols + 1), key=lambda i: qty_hits[i]))
    if any(sku_len):   candidates["article"] = int(max(range(1, max_cols + 1), key=lambda i: sku_len[i]))

    # customer_sku fallback same as article (user can change in UI)
    candidates["customer_sku"] = candidates["article"]

    return candidates

def _normalize_date(raw) -> Optional[datetime.date]:
    from dateutil import parser as date_parser
    try:
        if raw is None or raw == "":
            return None
        if isinstance(raw, (int, float)) and raw > 59:  # Excel serial
            # Excel's serial date origin 1899-12-30
            return (datetime(1899, 12, 30) + timedelta(days=float(raw))).date()
        return date_parser.parse(str(raw)).date()
    except Exception:
        return None

def _to_float(raw) -> Optional[float]:
    if raw is None or raw == "":
        return None
    try:
        s = str(raw).replace(",", "").replace(" ", "")
        s = s.replace("(", "-").replace(")", "")
        return float(s)
    except Exception:
        return None

def _parse_preview(fpath: str, mapping: Dict[str, Any], limit: int = 100) -> List[Dict[str, Any]]:
    """
    Returns list of parsed dicts for preview. Does NOT insert DB rows.
    """
    wb = _load_workbook(fpath)
    ws = wb[mapping["sheet"]]
    start = int(mapping.get("data_start_row") or 2)
    maxr = ws.max_row or start
    fields = mapping.get("fields", {})

    out: List[Dict[str, Any]] = []
    for r in range(start, min(maxr, start + limit - 1) + 1):
        rec: Dict[str, Any] = {"SourceRow": r}
        errs: List[str] = []

        def cell(col_idx):
            if not col_idx:
                return None
            return ws.cell(row=r, column=int(col_idx)).value

        # Date
        c_date = fields.get("date", {}).get("col")
        dt = _normalize_date(cell(c_date)) if c_date else None
        if c_date and dt is None:
            errs.append("Invalid date")
        rec["Date"] = dt

        # Article
        c_art = fields.get("article", {}).get("col")
        art = str(cell(c_art)).strip().upper() if c_art and cell(c_art) not in (None, "") else None
        rec["Article"] = art

        # CustomerSKU
        c_csku = fields.get("customer_sku", {}).get("col")
        csku = str(cell(c_csku)).strip() if c_csku and cell(c_csku) not in (None, "") else None
        rec["CustomerSKU"] = csku

        # Qty
        c_qty = fields.get("qty", {}).get("col")
        qty = _to_float(cell(c_qty)) if c_qty else None
        if c_qty and qty is None:
            errs.append("Invalid qty")
        rec["Qty"] = qty

        # Optionals
        for key in ("Store", "Region", "InvoiceNo", "Brand", "Site"):
            f = fields.get(key.lower(), {})
            col = f.get("col")
            val = str(cell(col)).strip() if col and cell(col) not in (None, "") else None
            rec[key] = val

        rec["ValidationErr"] = errs[0] if errs else None
        rec["ErrorsJSON"] = json.dumps(errs) if errs else None
        out.append(rec)
    return out

def _insert_staging(upload_id: int, customer_id: int, profile_id: Optional[int],
                    fpath: str, mapping: Dict[str, Any]) -> int:
    """
    Parse entire sheet and insert rows into SP_SellOut_Staging for an UploadID.
    Returns inserted rows count.
    """
    wb = _load_workbook(fpath)
    ws = wb[mapping["sheet"]]
    start = int(mapping.get("data_start_row") or 2)
    maxr = ws.max_row or start
    fields = mapping.get("fields", {})

    inserted = 0
    for r in range(start, maxr + 1):
        errs: List[str] = []

        def cell(col_idx):
            if not col_idx:
                return None
            return ws.cell(row=r, column=int(col_idx)).value

        dt = _normalize_date(cell(fields.get("date", {}).get("col")))
        art = None
        if fields.get("article", {}).get("col"):
            v = cell(fields["article"]["col"])
            art = str(v).strip().upper() if v not in (None, "") else None

        csku = None
        if fields.get("customer_sku", {}).get("col"):
            v = cell(fields["customer_sku"]["col"])
            csku = str(v).strip() if v not in (None, "") else None

        qty = _to_float(cell(fields.get("qty", {}).get("col")))

        if fields.get("date", {}).get("col") and dt is None:
            errs.append("Invalid date")
        if fields.get("qty", {}).get("col") and qty is None:
            errs.append("Invalid qty")

        vals = {
            "UploadID": upload_id,
            "ProfileID": profile_id,
            "CustomerID": customer_id,
            "SourceFileName": os.path.basename(fpath),
            "SourceSheet": mapping["sheet"],
            "SourceRow": r,
            "Date": dt,
            "Article": art,
            "CustomerSKU": csku,
            "Qty": qty,
            "Store": None,
            "Region": None,
            "InvoiceNo": None,
            "Brand": None,
            "Site": None,
            "ValidationErr": errs[0] if errs else None,
            "ErrorsJSON": json.dumps(errs) if errs else None,
            "CreatedAt": datetime.utcnow(),
        }

        # Optionals
        for key in ("store", "region", "invoice", "brand", "site"):
            if key in fields and fields[key].get("col"):
                v = cell(fields[key]["col"])
                vals[key.capitalize() if key != "invoice" else "InvoiceNo"] = (str(v).strip() if v not in (None, "") else None)

        row = SP_SellOut_Staging(**vals)
        model.add(row)
        inserted += 1

    model.commit()
    return inserted

# ---------------------------
# Routes
# ---------------------------

@bp.get("/")
# @login_required
def list_profiles():
    customer_id = request.args.get("customer_id", type=int)
    q = model.query(SP_CustomerUploadProfile).order_by(SP_CustomerUploadProfile.CustomerID, SP_CustomerUploadProfile.ProfileName)
    if customer_id:
        q = q.filter(SP_CustomerUploadProfile.CustomerID == customer_id)
    profiles = q.all()

    customers = model.query(SP_Customer).order_by(SP_Customer.CustName).all()
    return render_template("sellout_profiles/list.html", profiles=profiles, customers=customers, selected_customer_id=customer_id)


@bp.get("/create")
# @login_required
def create_form():
    customers = model.query(SP_Customer).order_by(SP_Customer.CustName).all()
    return render_template("sellout_profiles/form.html",
                           mode="create",
                           customers=customers,
                           suggested=None,
                           mapping=None)


@bp.post("/detect")
def detect():
    customer_id = request.form.get("CustomerID", type=int)
    file = request.files.get("file")
    if not file or not _allowed_file(file.filename):
        flash("Please upload a valid Excel file.", "danger")
        return redirect(url_for(".create_form"))

    token, fpath = _save_temp_upload(file)

    sheet, header_row, data_start = _guess_sheet_and_header(fpath)
    cols = _suggest_columns(fpath, sheet, header_row, data_start)

    suggested_mapping = {
        "sheet": sheet,
        "header_row": header_row,
        "data_start_row": data_start,
        "fields": {
            "date": {"col": cols.get("date")},
            "article": {"col": cols.get("article")},
            "qty": {"col": cols.get("qty")},
            "customer_sku": {"col": cols.get("customer_sku")},
        }
    }

    customers = model.query(SP_Customer).order_by(SP_Customer.CustName).all()
    return render_template(
        "sellout_profiles/form.html",
        mode="create",
        customers=customers,
        selected_customer_id=customer_id,
        token=token,
        suggested=suggested_mapping,
        mapping=json.dumps(suggested_mapping, ensure_ascii=False)
    )


@bp.post("/preview")
def preview():
    token = request.form.get("Token")
    mapping_raw = request.form.get("MappingJSON")
    if not token or not mapping_raw:
        flash("Missing token or mapping.", "danger")
        return redirect(url_for(".create_form"))

    fpath = _resolve_token_path(token)
    if not fpath or not os.path.exists(fpath):
        flash("Upload expired. Please re-upload.", "warning")
        return redirect(url_for(".create_form"))

    try:
        mapping = json.loads(mapping_raw)
    except Exception:
        flash("Invalid mapping JSON.", "danger")
        return redirect(url_for(".create_form"))

    rows = _parse_preview(fpath, mapping, limit=100)
    return render_template("sellout_profiles/preview.html",
                           mapping=mapping,
                           rows=rows,
                           filename=os.path.basename(fpath),
                           token=token)


@bp.post("/save")
# @login_required
def save_profile():
    """
    Persist profile header + detail JSON.
    """
    customer_id = request.form.get("CustomerID", type=int)
    profile_name = request.form.get("ProfileName", type=str)
    mapping_raw = request.form.get("MappingJSON")
    is_default = bool(request.form.get("IsDefault"))

    if not (customer_id and profile_name and mapping_raw):
        flash("Customer, Profile Name, and Mapping are required.", "danger")
        return redirect(url_for(".create_form"))

    try:
        mapping = json.loads(mapping_raw)
    except Exception:
        flash("Invalid mapping JSON.", "danger")
        return redirect(url_for(".create_form"))

    try:
        prof = SP_CustomerUploadProfile(
            CustomerID=customer_id,
            ProfileName=profile_name.strip(),
            IsActive=True,
            IsDefault=is_default,
            SheetName=mapping.get("sheet"),
            HeaderRowIndex=mapping.get("header_row"),
            DataStartRow=mapping.get("data_start_row"),
            Notes=None,
            CreatedBy="system",  # or current_user.username
        )
        model.add(prof)
        model.flush()  # get ProfileID

        detail = SP_CustomerUploadProfileDetail(
            ProfileID=prof.ProfileID,
            MappingJSON=json.dumps(mapping, ensure_ascii=False)
        )
        model.add(detail)

        if is_default:
            # unset other defaults for this customer
            model.query(SP_CustomerUploadProfile)\
                 .filter(and_(SP_CustomerUploadProfile.CustomerID == customer_id,
                              SP_CustomerUploadProfile.ProfileID != prof.ProfileID,
                              SP_CustomerUploadProfile.IsDefault == True))\
                 .update({SP_CustomerUploadProfile.IsDefault: False}, synchronize_session=False)

        model.commit()
        flash("Profile saved successfully.", "success")
        return redirect(url_for(".list_profiles", customer_id=customer_id))
    except SQLAlchemyError as e:
        model.rollback()
        flash(f"DB error while saving profile: {str(e)}", "danger")
        return redirect(url_for(".create_form"))


@bp.get("/<int:profile_id>/edit")
# @login_required
def edit_profile(profile_id: int):
    prof = model.query(SP_CustomerUploadProfile).get(profile_id)
    if not prof:
        flash("Profile not found.", "warning")
        return redirect(url_for(".list_profiles"))

    detail = model.query(SP_CustomerUploadProfileDetail).get(profile_id)
    customers = model.query(SP_Customer).order_by(SP_Customer.CustName).all()

    return render_template("sellout_profiles/form.html",
                           mode="edit",
                           profile=prof,
                           customers=customers,
                           selected_customer_id=prof.CustomerID,
                           mapping=(detail.MappingJSON if detail else "{}"))


@bp.post("/<int:profile_id>/update")
# @login_required
def update_profile(profile_id: int):
    prof = model.query(SP_CustomerUploadProfile).get(profile_id)
    if not prof:
        flash("Profile not found.", "warning")
        return redirect(url_for(".list_profiles"))

    customer_id = request.form.get("CustomerID", type=int) or prof.CustomerID
    profile_name = request.form.get("ProfileName", type=str) or prof.ProfileName
    mapping_raw = request.form.get("MappingJSON")
    is_default = bool(request.form.get("IsDefault"))

    try:
        mapping = json.loads(mapping_raw) if mapping_raw else {}
    except Exception:
        flash("Invalid mapping JSON.", "danger")
        return redirect(url_for(".edit_profile", profile_id=profile_id))

    try:
        prof.CustomerID = customer_id
        prof.ProfileName = profile_name.strip()
        prof.IsDefault = is_default
        prof.SheetName = mapping.get("sheet")
        prof.HeaderRowIndex = mapping.get("header_row")
        prof.DataStartRow = mapping.get("data_start_row")
        prof.UpdatedAt = datetime.utcnow()
        prof.UpdatedBy = "system"

        det = model.query(SP_CustomerUploadProfileDetail).get(profile_id)
        if not det:
            det = SP_CustomerUploadProfileDetail(ProfileID=profile_id, MappingJSON=json.dumps(mapping, ensure_ascii=False))
            model.add(det)
        else:
            det.MappingJSON = json.dumps(mapping, ensure_ascii=False)

        if is_default:
            model.query(SP_CustomerUploadProfile)\
                 .filter(and_(SP_CustomerUploadProfile.CustomerID == customer_id,
                              SP_CustomerUploadProfile.ProfileID != profile_id,
                              SP_CustomerUploadProfile.IsDefault == True))\
                 .update({SP_CustomerUploadProfile.IsDefault: False}, synchronize_session=False)

        model.commit()
        flash("Profile updated.", "success")
        return redirect(url_for(".list_profiles", customer_id=customer_id))
    except SQLAlchemyError as e:
        model.rollback()
        flash(f"DB error while updating profile: {str(e)}", "danger")
        return redirect(url_for(".edit_profile", profile_id=profile_id))


@bp.post("/<int:profile_id>/toggle-active")
# @login_required
def toggle_active(profile_id: int):
    prof = model.query(SP_CustomerUploadProfile).get(profile_id)
    if not prof:
        flash("Profile not found.", "warning")
        return redirect(url_for(".list_profiles"))
    prof.IsActive = not bool(prof.IsActive)
    prof.UpdatedAt = datetime.utcnow()
    model.commit()
    flash(("Activated" if prof.IsActive else "Deactivated") + " profile.", "info")
    return redirect(url_for(".list_profiles", customer_id=prof.CustomerID))


@bp.post("/<int:profile_id>/set-default")
# @login_required
def set_default(profile_id: int):
    prof = model.query(SP_CustomerUploadProfile).get(profile_id)
    if not prof:
        flash("Profile not found.", "warning")
        return redirect(url_for(".list_profiles"))

    prof.IsDefault = True
    model.query(SP_CustomerUploadProfile)\
         .filter(and_(SP_CustomerUploadProfile.CustomerID == prof.CustomerID,
                      SP_CustomerUploadProfile.ProfileID != profile_id,
                      SP_CustomerUploadProfile.IsDefault == True))\
         .update({SP_CustomerUploadProfile.IsDefault: False}, synchronize_session=False)
    model.commit()
    flash("Set as default for this customer.", "success")
    return redirect(url_for(".list_profiles", customer_id=prof.CustomerID))


@bp.get("/use")
# @login_required
def use_form():
    """
    Simple page to select a profile + upload a file to stage.
    """
    customers = model.query(SP_Customer).order_by(SP_Customer.CustName).all()
    profiles = model.query(SP_CustomerUploadProfile).filter_by(IsActive=True).order_by(
        SP_CustomerUploadProfile.CustomerID, SP_CustomerUploadProfile.IsDefault.desc(), SP_CustomerUploadProfile.ProfileName
    ).all()
    return render_template("sellout_profiles/use.html", customers=customers, profiles=profiles)


@bp.post("/use")
# @login_required
def use_profile():
    """
    Create a SellOut Upload batch, parse entire file with selected profile, and land rows in staging.
    """
    profile_id = request.form.get("ProfileID", type=int)
    document_date = request.form.get("DocumentDate")  # yyyy-mm-dd (legacy field on header)
    upload_type = request.form.get("UploadType", default="Transactional")
    level_type = request.form.get("LevelType", default="HO")
    brand = request.form.get("Brand")
    period_start = request.form.get("PeriodStart")
    period_end = request.form.get("PeriodEnd")

    file = request.files.get("file")
    if not profile_id or not file or not _allowed_file(file.filename):
        flash("Profile and a valid Excel file are required.", "danger")
        return redirect(url_for(".use_form"))

    prof = model.query(SP_CustomerUploadProfile).get(profile_id)
    if not prof or not prof.IsActive:
        flash("Selected profile not found or inactive.", "warning")
        return redirect(url_for(".use_form"))

    det = model.query(SP_CustomerUploadProfileDetail).get(profile_id)
    if not det:
        flash("Profile mapping JSON is missing.", "danger")
        return redirect(url_for(".use_form"))

    mapping = json.loads(det.MappingJSON)

    with tempfile.TemporaryDirectory() as td:
        fname = secure_filename(file.filename)
        fpath = os.path.join(td, fname)
        file.save(fpath)
        fhash = _file_hash(fpath)

        # Create header row in SP_SellOutUploads (status Draft)
        hdr = SP_SellOutUploads(
            CustomerID=prof.CustomerID,
            LevelType=level_type,
            UploadType=upload_type,
            Brand=brand,
            DocumentDate=datetime.strptime(document_date, "%Y-%m-%d").date() if document_date else datetime.utcnow().date(),
            PeriodStart=datetime.strptime(period_start, "%Y-%m-%d").date() if period_start else None,
            PeriodEnd=datetime.strptime(period_end, "%Y-%m-%d").date() if period_end else None,
            Status="Draft",
            CreatedBy="system",
            CreatedAt=datetime.utcnow(),
            SourceFileName=fname,
            SourceFileHash=fhash,
            Notes=f"Uploaded via profile {prof.ProfileName}"
        )
        model.add(hdr)
        model.flush()  # get UploadID

        # Insert staging rows
        inserted = _insert_staging(
            upload_id=hdr.UploadID,
            customer_id=prof.CustomerID,
            profile_id=prof.ProfileID,
            fpath=fpath,
            mapping=mapping
        )

        model.commit()
        flash(f"File parsed and {inserted} rows staged (UploadID={hdr.UploadID}).", "success")
        return redirect(url_for(".view_staging", upload_id=hdr.UploadID))


@bp.get("/upload/<int:upload_id>/staging")
# @login_required
def view_staging(upload_id: int):
    rows = (model.query(SP_SellOut_Staging)
                 .filter(SP_SellOut_Staging.UploadID == upload_id)
                 .order_by(SP_SellOut_Staging.SourceRow)
                 .limit(5000)
                 .all())
    return render_template("sellout_profiles/staging.html", upload_id=upload_id, rows=rows)



@bp.post("/introspect")
def introspect():
    """
    Body: { token, sheet, header_row }
    Returns sheets, columns (A..), and a tiny sample of rows  (optional).
    """
    data = request.get_json(force=True)
    token = data.get("token")
    sheet = data.get("sheet")
    header_row = int(data.get("header_row") or 1)

    fpath = _resolve_token_path(token)
    if not fpath or not os.path.exists(fpath):
        return jsonify({"ok": False, "error": "Upload expired. Re-upload the file."}), 400

    wb = _load_workbook(fpath)
    sheets = wb.sheetnames
    if sheet not in sheets:
        sheet = sheets[0]

    cols = _sheet_headers(fpath, sheet, header_row)

    # light sample (first 15 rows after header)
    ws = wb[sheet]
    start = header_row + 1
    end = min(ws.max_row or start, start + 14)
    sample = []
    for r in range(start, end + 1):
        row_vals = []
        for c in range(1, len(cols) + 1):
            v = ws.cell(row=r, column=c).value
            row_vals.append("" if v is None else str(v))
        sample.append({"r": r, "vals": row_vals})

    return jsonify({"ok": True, "sheets": sheets, "sheet": sheet, "columns": cols, "sample": sample})
