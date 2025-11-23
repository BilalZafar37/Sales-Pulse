# masters_bp.py
from flask import Blueprint, request, render_template, jsonify, send_file
from sqlalchemy import or_, func
from models import (
    model, SP_Customer, SP_SKU, SP_Customer_SKU_Map, SP_CategoriesMappingMain
)
from config import STATIC_DIR, BASE_DIR

bp = Blueprint("masters", __name__, static_folder=STATIC_DIR, url_prefix="/masters")

@bp.route("/", methods=["GET"])
def masters_ui():
    return render_template("./masters/index.html")  # single page with tabs

# ---------- Helpers ----------
def _qparam(name, cast=str, default=None):
    v = request.args.get(name)
    return cast(v) if v is not None and v != "" else default

def _paginate(query):
    page      = _qparam("page", int, 1)
    page_size = min(100, _qparam("page_size", int, 20))
    total = query.count()
    rows  = query.order_by().offset((page-1)*page_size).limit(page_size).all()
    return total, rows

def _cat_json(c):
    if not c: return None
    return {
        "ID": c.ID,
        "CatCode": c.CatCode,
        "CatName": c.CatName,
        "CatDesc": c.CatDesc,
        "SubCat": c.SubCat
    }

# ---------- Autocomplete (Select2) ----------
@bp.route("/api/options/customers")
def opt_customers():
    term = (request.args.get("term") or "").strip()
    q = model.query(SP_Customer)
    if term:
        like = f"%{term}%"
        q = q.filter(or_(SP_Customer.CustName.ilike(like), SP_Customer.CustCode.ilike(like)))
    rows = q.order_by(SP_Customer.CustName).all()
    # If you prefer Name(Code), switch to: f"{c.CustName} ({c.CustCode})"
    return jsonify(results=[{"id": c.CustomerID, "text": c.CustName} for c in rows])

@bp.route("/api/options/categories")
def opt_categories():
    term = (request.args.get("term") or "").strip()
    q = model.query(SP_CategoriesMappingMain)
    if term:
        like = f"%{term}%"
        q = q.filter(or_(
            SP_CategoriesMappingMain.CatCode.ilike(like),
            SP_CategoriesMappingMain.CatName.ilike(like),
            SP_CategoriesMappingMain.CatDesc.ilike(like),
            SP_CategoriesMappingMain.SubCat.ilike(like),
        ))
    rows = q.order_by(SP_CategoriesMappingMain.CatCode).limit(50).all()
    return jsonify(results=[{
        "id": r.ID,
        "text": f"{r.CatCode} — {r.CatName or ''}"
    } for r in rows])

@bp.route("/api/options/skus")
def opt_skus():
    term    = (request.args.get("term") or "").strip()
    brand   = (request.args.get("brand") or "").strip()
    cat_id  = request.args.get("category_id", type=int)
    catcode = (request.args.get("catcode") or "").strip()

    q = model.query(SP_SKU)
    if brand:
        q = q.filter(SP_SKU.Brand == brand)
    if cat_id:
        q = q.filter(SP_SKU.CategoryMappingID == cat_id)
    elif catcode:
        q = q.join(SP_CategoriesMappingMain, SP_CategoriesMappingMain.ID == SP_SKU.CategoryMappingID)\
             .filter(SP_CategoriesMappingMain.CatCode == catcode)

    if term:
        like = f"%{term}%"
        q = q.filter(or_(SP_SKU.ArticleCode.ilike(like), SP_SKU.Description.ilike(like)))

    rows = q.order_by(SP_SKU.ArticleCode).limit(50).all()
    return jsonify(results=[{"id": s.SKU_ID, "text": f"{s.ArticleCode} — {s.Description or ''}"} for s in rows])

# ---------- Categories CRUD ----------
@bp.route("/api/categories", methods=["GET"])
def categories_list():
    """
    List categories with optional search over CatCode/CatName/CatDesc/SubCat
    """
    term = (request.args.get("q") or "").strip()
    q = model.query(SP_CategoriesMappingMain)
    if term:
        like = f"%{term}%"
        q = q.filter(or_(
            SP_CategoriesMappingMain.CatCode.ilike(like),
            SP_CategoriesMappingMain.CatName.ilike(like),
            SP_CategoriesMappingMain.CatDesc.ilike(like),
            SP_CategoriesMappingMain.SubCat.ilike(like),
        ))

    # (optional) include a count of SKUs referencing each category
    q = q.add_columns(
        (model.query(func.count(SP_SKU.SKU_ID))
            .filter(SP_SKU.CategoryMappingID == SP_CategoriesMappingMain.ID)
            .correlate(SP_CategoriesMappingMain)
            .as_scalar()
        ).label("SKUCount")
    ).order_by(SP_CategoriesMappingMain.CatCode)

    # paginate
    page      = _qparam("page", int, 1)
    page_size = min(100, _qparam("page_size", int, 50))
    total = q.count()
    rows = q.offset((page-1)*page_size).limit(page_size).all()

    items = []
    for c, sku_cnt in rows:
        items.append({
            **_cat_json(c),
            "SKUCount": int(sku_cnt or 0)
        })
    return jsonify(ok=True, total=total, items=items)

@bp.route("/api/categories", methods=["POST"])
def categories_create():
    data = request.get_json(force=True)
    code = (data.get("CatCode") or "").strip().upper()
    name = (data.get("CatName") or "").strip() or None
    desc = (data.get("CatDesc") or "").strip() or None
    subc = (data.get("SubCat")  or "").strip() or None

    if not code:
        return jsonify(ok=False, error="CatCode is required"), 400
    # If you want to enforce 3 letters:
    # if len(code) != 3: return jsonify(ok=False, error="CatCode must be 3 characters"), 400

    exists = model.query(SP_CategoriesMappingMain).filter(SP_CategoriesMappingMain.CatCode == code).first()
    if exists:
        return jsonify(ok=False, error="CatCode already exists"), 409

    obj = SP_CategoriesMappingMain(CatCode=code, CatName=name, CatDesc=desc, SubCat=subc)
    model.add(obj); model.commit()
    return jsonify(ok=True, id=obj.ID)

@bp.route("/api/categories/<int:cid>", methods=["PATCH"])
def categories_update(cid):
    data = request.get_json(force=True)
    obj = model.query(SP_CategoriesMappingMain).get(cid)
    if not obj:
        return jsonify(ok=False, error="Not found"), 404

    if "CatCode" in data:
        new_code = (data["CatCode"] or "").strip().upper()
        if not new_code:
            return jsonify(ok=False, error="CatCode cannot be empty"), 400
        # if len(new_code) != 3: return jsonify(ok=False, error="CatCode must be 3 characters"), 400
        dup = (model.query(SP_CategoriesMappingMain)
                     .filter(SP_CategoriesMappingMain.CatCode == new_code,
                             SP_CategoriesMappingMain.ID != cid).first())
        if dup:
            return jsonify(ok=False, error="CatCode already exists"), 409
        obj.CatCode = new_code

    for field in ("CatName", "CatDesc", "SubCat"):
        if field in data:
            val = (data[field] or "").strip()
            setattr(obj, field, val if val else None)

    model.commit()
    return jsonify(ok=True)

@bp.route("/api/categories/<int:cid>", methods=["DELETE"])
def categories_delete(cid):
    """
    By default, refuses delete if any SKUs reference this category.
    Pass ?force=1 to null CategoryMappingID on those SKUs, then delete.
    """
    force = request.args.get("force", "0") in ("1", "true", "yes")
    obj = model.query(SP_CategoriesMappingMain).get(cid)
    if not obj:
        return jsonify(ok=False, error="Not found"), 404

    ref_count = model.query(func.count(SP_SKU.SKU_ID)).filter(SP_SKU.CategoryMappingID == cid).scalar() or 0
    if ref_count and not force:
        return jsonify(ok=False, error=f"Category in use by {ref_count} SKU(s). Use ?force=1 to detach then delete."), 409

    if ref_count and force:
        # Null out references
        model.query(SP_SKU).filter(SP_SKU.CategoryMappingID == cid)\
             .update({SP_SKU.CategoryMappingID: None}, synchronize_session=False)

    model.delete(obj); model.commit()
    return jsonify(ok=True)

# ---------- Customers CRUD ----------
@bp.route("/api/customers", methods=["GET"])
def customers_list():
    q = model.query(SP_Customer)
    term = (request.args.get("q") or "").strip()
    if term:
        like = f"%{term}%"
        q = q.filter(or_(SP_Customer.CustName.ilike(like), SP_Customer.CustCode.ilike(like)))
    total, rows = _paginate(q.order_by(SP_Customer.CustName))
    
    # Getting the actual parent code 
    # Build a map {ParentCustomerID -> ParentCustCode} for the current page
    parent_ids = {r.ParentCustID for r in rows if r.ParentCustID}
    parent_map = {}
    if parent_ids:
        pairs = model.query(SP_Customer.CustomerID, SP_Customer.CustCode)\
                     .filter(SP_Customer.CustomerID.in_(parent_ids))\
                     .all()
        parent_map = {cid: code for cid, code in pairs}
    
    items = [{
        "CustomerID": r.CustomerID,
        "CustCode": r.CustCode,
        "CustName": r.CustName,
        "LevelType": r.LevelType,
        "ParentCustID": parent_map.get(r.ParentCustID), # get actual code from map (ignore the wrong key name)
    } for r in rows]
    return jsonify(ok=True, total=total, items=items)

@bp.route("/api/customers", methods=["POST"])
def customers_create():
    data = request.get_json(force=True)
    c = SP_Customer(
        CustCode=data["CustCode"].strip(),
        CustName=data["CustName"].strip(),
        LevelType=data.get("LevelType","HO"),
        ParentCustID=data.get("ParentCustID")
    )
    model.add(c); model.commit()
    return jsonify(ok=True, id=c.CustomerID)

@bp.route("/api/customers/<int:cid>", methods=["PATCH"])
def customers_update(cid):
    data = request.get_json(force=True)
    c = model.query(SP_Customer).get(cid)
    if not c: return jsonify(ok=False, error="Not found"), 404
    for k in ("CustCode","CustName","LevelType","ParentCustID"):
        if k in data: setattr(c, k, data[k])
    model.commit()
    return jsonify(ok=True)

@bp.route("/api/customers/<int:cid>", methods=["DELETE"])
def customers_delete(cid):
    c = model.query(SP_Customer).get(cid)
    if not c: return jsonify(ok=False, error="Not found"), 404
    model.delete(c); model.commit()
    return jsonify(ok=True)

# ---------- SKUs CRUD (uses CategoryMappingID) ----------
@bp.route("/api/skus", methods=["GET"])
def skus_list():
    term    = (request.args.get("q") or "").strip()
    brand   = (request.args.get("brand") or "").strip()
    cat_id  = request.args.get("category_id", type=int)
    catcode = (request.args.get("catcode") or "").strip()

    q = model.query(SP_SKU, SP_CategoriesMappingMain)\
             .outerjoin(SP_CategoriesMappingMain, SP_CategoriesMappingMain.ID == SP_SKU.CategoryMappingID)

    if term:
        like = f"%{term}%"
        q = q.filter(or_(SP_SKU.ArticleCode.ilike(like), SP_SKU.Description.ilike(like)))
    if brand:
        q = q.filter(SP_SKU.Brand == brand)
    if cat_id:
        q = q.filter(SP_SKU.CategoryMappingID == cat_id)
    elif catcode:
        q = q.filter(SP_CategoriesMappingMain.CatCode == catcode)

    total = q.count()
    page      = _qparam("page", int, 1)
    page_size = min(100, _qparam("page_size", int, 20))
    rows = (q.order_by(SP_SKU.ArticleCode)
              .offset((page-1)*page_size).limit(page_size).all())

    items = []
    for s, c in rows:
        items.append({
            "SKU_ID": s.SKU_ID,
            "ArticleCode": s.ArticleCode,
            "Description": s.Description,
            "Brand": s.Brand,
            "CategoryMappingID": s.CategoryMappingID,
            "Category": _cat_json(c)
        })
    return jsonify(ok=True, total=total, items=items)

@bp.route("/api/skus", methods=["POST"])
def skus_create():
    data = request.get_json(force=True)
    # resolve category
    cat_id  = data.get("CategoryMappingID")
    catcode = (data.get("CatCode") or data.get("catcode") or "").strip().upper()
    if not cat_id and catcode:
        m = model.query(SP_CategoriesMappingMain).filter(SP_CategoriesMappingMain.CatCode == catcode).first()
        cat_id = m.ID if m else None

    s = SP_SKU(
        ArticleCode       = data["ArticleCode"].strip(),
        Description       = (data.get("Description") or None),
        Brand             = (data.get("Brand") or None),
        CategoryMappingID = cat_id
    )
    model.add(s); model.commit()
    return jsonify(ok=True, id=s.SKU_ID)

@bp.route("/api/skus/<int:sku_id>", methods=["PATCH"])
def skus_update(sku_id):
    s = model.query(SP_SKU).get(sku_id)
    if not s: return jsonify(ok=False, error="Not found"), 404
    data = request.get_json(force=True)

    for k in ("ArticleCode","Description","Brand"):
        if k in data:
            v = data[k]
            setattr(s, k, v.strip() if isinstance(v, str) else v)

    if "CategoryMappingID" in data or "CatCode" in data or "catcode" in data:
        cat_id  = data.get("CategoryMappingID")
        catcode = (data.get("CatCode") or data.get("catcode") or "").strip().upper()
        if not cat_id and catcode:
            m = model.query(SP_CategoriesMappingMain).filter(SP_CategoriesMappingMain.CatCode == catcode).first()
            cat_id = m.ID if m else None
        s.CategoryMappingID = cat_id

    model.commit()
    return jsonify(ok=True)

@bp.route("/api/skus/<int:sku_id>", methods=["DELETE"])
def skus_delete(sku_id):
    s = model.query(SP_SKU).get(sku_id)
    if not s: return jsonify(ok=False, error="Not found"), 404
    model.delete(s); model.commit()
    return jsonify(ok=True)

# ---------- Mapping CRUD ----------
@bp.route("/api/mappings", methods=["GET"])
def mappings_list():
    cust_id = request.args.get("customer_id", type=int)

    q = (model.query(SP_Customer_SKU_Map, SP_SKU, SP_CategoriesMappingMain, SP_Customer)
         .join(SP_SKU, SP_SKU.SKU_ID == SP_Customer_SKU_Map.SKU_ID)
         .outerjoin(SP_CategoriesMappingMain, SP_CategoriesMappingMain.ID == SP_SKU.CategoryMappingID)
         .join(SP_Customer, SP_Customer.CustomerID == SP_Customer_SKU_Map.CustomerID))

    if cust_id:
        q = q.filter(SP_Customer_SKU_Map.CustomerID == cust_id)

    total = q.count()
    page      = _qparam("page", int, 1)
    page_size = min(100, _qparam("page_size", int, 100))
    rows = (q.order_by(SP_Customer_SKU_Map.MapID.desc())
              .offset((page-1)*page_size).limit(page_size).all())

    items = []
    for m, s, c, cust in rows:
        items.append({
            "MapID": m.MapID,
            "CustomerID": m.CustomerID,
            "Customer": f"{cust.CustName} ({cust.CustCode})",
            "SKU_ID": m.SKU_ID,
            "ArticleCode": s.ArticleCode,
            "SKU_Desc": s.Description,
            "Brand": s.Brand,
            "CategoryMappingID": s.CategoryMappingID,
            "Category": _cat_json(c),
            "CustSKUCode": m.CustSKUCode,
        })
    return jsonify(ok=True, total=total, items=items)

@bp.route("/api/mappings", methods=["POST"])
def mappings_create():
    data = request.get_json(force=True)
    m = SP_Customer_SKU_Map(
        CustomerID = data["CustomerID"],
        SKU_ID     = data["SKU_ID"],
        CustSKUCode= data["CustSKUCode"].strip()
    )
    model.add(m); model.commit()
    return jsonify(ok=True, id=m.MapID)

@bp.route("/api/mappings/<int:map_id>", methods=["PATCH"])
def mappings_update(map_id):
    m = model.query(SP_Customer_SKU_Map).get(map_id)
    if not m: return jsonify(ok=False, error="Not found"), 404
    data = request.get_json(force=True)
    for k in ("CustomerID","SKU_ID","CustSKUCode"):
        if k in data:
            v = data[k]
            if k == "CustSKUCode" and isinstance(v, str):
                v = v.strip()
            setattr(m, k, v)
    model.commit()
    return jsonify(ok=True)

@bp.route("/api/mappings/<int:map_id>", methods=["DELETE"])
def mappings_delete(map_id):
    m = model.query(SP_Customer_SKU_Map).get(map_id)
    if not m: return jsonify(ok=False, error="Not found"), 404
    model.delete(m); model.commit()
    return jsonify(ok=True)

# ---------- Bulk Paste for mapping ----------
@bp.route("/api/mappings/paste", methods=["POST"])
def mappings_paste():
    """
    JSON: { "CustomerID": 123, "rows": [ {"CustSKUCode":"A1","ArticleCode":"X001"}, ... ] }
    """
    data = request.get_json(force=True)
    cust_id = int(data["CustomerID"])
    rows    = data.get("rows", [])

    codes = [ (r.get("ArticleCode") or "").strip() for r in rows if r.get("ArticleCode") ]
    if not codes:
        return jsonify(ok=False, error="No rows"), 400

    skus = model.query(SP_SKU).filter(SP_SKU.ArticleCode.in_(codes)).all()
    code_to_id = {s.ArticleCode: s.SKU_ID for s in skus}

    created, skipped = [], []
    for r in rows:
        art  = (r.get("ArticleCode") or "").strip()
        csku = (r.get("CustSKUCode") or "").strip()
        sku_id = code_to_id.get(art)
        if not sku_id or not csku:
            skipped.append({"ArticleCode": art, "reason": "missing/unknown"})
            continue
        exists = model.query(SP_Customer_SKU_Map)\
            .filter_by(CustomerID=cust_id, SKU_ID=sku_id).first()
        if exists:
            exists.CustSKUCode = csku
        else:
            model.add(SP_Customer_SKU_Map(CustomerID=cust_id, SKU_ID=sku_id, CustSKUCode=csku))
            created.append(art)
    model.commit()
    return jsonify(ok=True, created=len(created), skipped=skipped)


# ===== Bulk Excel Templates + Upload (Customers, SKUs, Categories, Mappings) =====
from openpyxl import Workbook, load_workbook
from werkzeug.utils import secure_filename
import io

# --- Spec for each master: column order, required flags, notes ---
MASTER_SPECS = {
    "customers": {
        "sheet": "Customers",
        "key": "CustCode",  # natural key for upsert
        "columns": [
            ("CustCode",   True,  "Unique customer code"),
            ("CustName",   True,  "Customer name"),
            ("LevelType",  True,  "HO or Branch"),
            ("ParentCode", False, "Parent customer's CustCode, optional"),
        ],
        "sample": [
            {"CustCode": "CUST1001", "CustName": "ABC Trading", "LevelType": "HO", "ParentCode": ""},
            {"CustCode": "CUST1002", "CustName": "ABC Retail Riyadh", "LevelType": "Branch", "ParentCode": "CUST1001"},
        ],
    },
    "skus": {
        "sheet": "SKUs",
        "key": "ArticleCode",
        "columns": [
            ("ArticleCode",      True,  "Unique article code"),
            ("Description",      False, "Short description"),
            ("Brand",            False, "Brand name"),
            ("CategoryCatCode",  False, "Category CatCode (optional)"),
        ],
        "sample": [
            {"ArticleCode":"ART-1001", "Description":"Widget S", "Brand":"ACME", "CategoryCatCode":"ELEC"},
            {"ArticleCode":"ART-1002", "Description":"Widget L", "Brand":"ACME", "CategoryCatCode":""},
        ],
    },
    "categories": {
        "sheet": "Categories",
        "key": "CatCode",
        "columns": [
            ("CatCode", True,  "Unique category code"),
            ("CatName", True,  "Category name"),
            ("CatDesc", False, "Description"),
            ("SubCat",  False, "Sub category"),
        ],
        "sample": [
            {"CatCode":"ELEC","CatName":"Electronics","CatDesc":"Small electronics","SubCat":"Gadgets"},
            {"CatCode":"HOME","CatName":"Home","CatDesc":"","SubCat":""},
        ],
    },
    "mappings": {
        "sheet": "CustSKUMap",
        "key": None,  # composite
        "columns": [
            ("CustCode",    True, "Customer's CustCode"),
            ("ArticleCode", True, "SKU ArticleCode"),
            ("CustSKUCode", True, "Customer's SKU code/alias"),
        ],
        "sample": [
            {"CustCode":"CUST1002","ArticleCode":"ART-1001","CustSKUCode":"WID-S"},
            {"CustCode":"CUST1002","ArticleCode":"ART-1002","CustSKUCode":"WID-L"},
        ],
    },
}

def _make_template_wb(master: str) -> bytes:
    spec = MASTER_SPECS[master]
    wb = Workbook()
    ws = wb.active
    ws.title = spec["sheet"]

    # Header with required mark
    headers = []
    for col, req, note in spec["columns"]:
        headers.append(f"{col}{' *' if req else ''}")
    ws.append(headers)

    # Notes row (italic)
    notes = [note for _, _, note in spec["columns"]]
    ws.append(notes)

    # Sample rows
    for samp in spec["sample"]:
        row = []
        for col, *_ in spec["columns"]:
            row.append(samp.get(col, ""))
        ws.append(row)

    # basic widths
    for i, (col, *_rest) in enumerate(spec["columns"], start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(14, len(col) + 2)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()

@bp.route("/api/bulk/template/<master>", methods=["GET"])
def bulk_template(master):
    master = master.lower()
    if master not in MASTER_SPECS:
        return jsonify(ok=False, error="Unknown master"), 400
    content = _make_template_wb(master)
    fname = f"{MASTER_SPECS[master]['sheet']}_Template.xlsx"
    return send_file(io.BytesIO(content), as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def _ws_to_dicts(ws, spec):
    """
    Parse a worksheet into list[dict]. Skip the first 2 rows (header + notes).
    Header names are taken from spec['columns'] (strict by position).
    """
    col_names = [c for c, *_ in spec["columns"]]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i <= 2:  # skip header + notes
            continue
        if all((v is None or str(v).strip() == "") for v in row):
            continue
        item = {}
        for j, name in enumerate(col_names):
            item[name] = (row[j] if j < len(row) else None)
            if isinstance(item[name], str):
                item[name] = item[name].strip()
        rows.append(item)
    return rows

@bp.route("/api/bulk/upload/<master>", methods=["POST"])
def bulk_upload(master):
    """
    Upload Excel (first sheet) and upsert rows for the selected master.
    Form-data: file=<xlsx>, validate_only=0/1 (optional)
    """
    master = master.lower()
    if master not in MASTER_SPECS:
        return jsonify(ok=False, error="Unknown master"), 400
    f = request.files.get("file")
    if not f:
        return jsonify(ok=False, error="No file"), 400
    try:
        wb = load_workbook(filename=io.BytesIO(f.read()), data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify(ok=False, error=f"Invalid Excel: {e}"), 400

    spec = MASTER_SPECS[master]
    rows = _ws_to_dicts(ws, spec)
    validate_only = (request.form.get("validate_only") in ("1", "true", "yes"))

    # route to specific handler
    if master == "customers":
        result = _bulk_upsert_customers(rows, validate_only)
    elif master == "skus":
        result = _bulk_upsert_skus(rows, validate_only)
    elif master == "categories":
        result = _bulk_upsert_categories(rows, validate_only)
    elif master == "mappings":
        result = _bulk_upsert_mappings(rows, validate_only)
    else:
        return jsonify(ok=False, error="Not implemented"), 400

    return jsonify(result)

# ----------------- per-master upsert handlers -----------------

def _bulk_upsert_categories(rows, validate_only: bool):
    required = {"CatCode","CatName"}
    created = updated = skipped = 0
    errors = []
    # prefetch existing by CatCode
    existing = {c.CatCode.upper(): c for c in model.query(SP_CategoriesMappingMain).all()}
    for i, r in enumerate(rows, start=1):
        # validate
        if any(not r.get(k) for k in required):
            skipped += 1
            errors.append({"row": i, "error": "Missing required CatCode or CatName"})
            continue
        code = (r["CatCode"] or "").strip().upper()
        name = (r["CatName"] or "").strip()
        desc = (r.get("CatDesc") or None)
        subc = (r.get("SubCat") or None)

        obj = existing.get(code)
        if obj:
            obj.CatName = name
            obj.CatDesc = desc
            obj.SubCat  = subc
            updated += 1
        else:
            obj = SP_CategoriesMappingMain(CatCode=code, CatName=name, CatDesc=desc, SubCat=subc, Brand=r.get("Brand"))
            model.add(obj)
            existing[code] = obj
            created += 1
    if not validate_only:
        model.commit()
    else:
        model.rollback()
    return {"ok": True, "master":"categories", "created":created, "updated":updated, "skipped":skipped, "errors":errors}

def _bulk_upsert_customers(rows, validate_only: bool):
    required = {"CustCode","CustName","LevelType"}
    created = updated = skipped = 0
    errors = []

    # cache existing customers by code + id map for parent resolution
    customers = {c.CustCode.upper(): c for c in model.query(SP_Customer).all()}

    for i, r in enumerate(rows, start=1):
        if any(not r.get(k) for k in required):
            skipped += 1
            errors.append({"row": i, "error": "Missing CustCode/CustName/LevelType"})
            continue
        code = r["CustCode"].strip().upper()
        name = r["CustName"].strip()
        level= (r["LevelType"] or "HO").strip()
        parent_code = (r.get("ParentCode") or "").strip().upper()
        parent_id = customers.get(parent_code).CustomerID if parent_code and customers.get(parent_code) else None

        obj = customers.get(code)
        if obj:
            obj.CustName = name
            obj.LevelType= level
            obj.ParentCustID = parent_id
            updated += 1
        else:
            obj = SP_Customer(CustCode=code, CustName=name, LevelType=level, ParentCustID=parent_id)
            model.add(obj)
            # flush to get ID so subsequent rows can reference as parent
            if not validate_only:
                model.commit()
            customers[code] = obj
            created += 1
    if not validate_only:
        model.commit()
    else:
        model.rollback()
    return {"ok": True, "master":"customers", "created":created, "updated":updated, "skipped":skipped, "errors":errors}

def _bulk_upsert_skus(rows, validate_only: bool):
    required = {"ArticleCode"}
    created = updated = skipped = 0
    errors = []

    # prefetch cats by CatCode for mapping
    cats = {c.CatCode.upper(): c.ID for c in model.query(SP_CategoriesMappingMain).all()}
    skus = {s.ArticleCode.upper(): s for s in model.query(SP_SKU).all()}

    for i, r in enumerate(rows, start=1):
        if any(not r.get(k) for k in required):
            skipped += 1
            errors.append({"row": i, "error": "Missing ArticleCode"})
            continue
        art  = r["ArticleCode"].strip().upper()
        desc = (r.get("Description") or None)
        brand= (r.get("Brand") or None)
        cat_code = (r.get("CategoryCatCode") or "").strip().upper()
        cat_id = cats.get(cat_code) if cat_code else None

        obj = skus.get(art)
        if obj:
            obj.Description = desc
            obj.Brand = brand
            obj.CategoryMappingID = cat_id
            updated += 1
        else:
            obj = SP_SKU(ArticleCode=art, Description=desc, Brand=brand, CategoryMappingID=cat_id)
            model.add(obj)
            skus[art] = obj
            created += 1
    if not validate_only:
        model.commit()
    else:
        model.rollback()
    return {"ok": True, "master":"skus", "created":created, "updated":updated, "skipped":skipped, "errors":errors}

def _bulk_upsert_mappings(rows, validate_only: bool):
    required = {"CustCode","ArticleCode","CustSKUCode"}
    created = updated = skipped = 0
    errors = []

    # prefetch maps: customer by code, sku by article
    cust_by_code = {c.CustCode.upper(): c.CustomerID for c in model.query(SP_Customer).all()}
    sku_by_code  = {s.ArticleCode.upper(): s.SKU_ID for s in model.query(SP_SKU).all()}

    # cache of existing mapping tuples
    existing = {(m.CustomerID, m.SKU_ID): m for m in model.query(SP_Customer_SKU_Map).all()}

    for i, r in enumerate(rows, start=1):
        if any(not r.get(k) for k in required):
            skipped += 1
            errors.append({"row": i, "error": "Missing CustCode/ArticleCode/CustSKUCode"})
            continue
        cust_code = r["CustCode"].strip().upper()
        art_code  = r["ArticleCode"].strip().upper()
        cust_sku  = r["CustSKUCode"].strip()

        cust_id = cust_by_code.get(cust_code)
        sku_id  = sku_by_code.get(art_code)

        if not cust_id or not sku_id:
            skipped += 1
            errors.append({"row": i, "error": f"Unknown CustCode or ArticleCode ({cust_code}, {art_code})"})
            continue

        key = (cust_id, sku_id)
        if key in existing:
            existing[key].CustSKUCode = cust_sku
            updated += 1
        else:
            model.add(SP_Customer_SKU_Map(CustomerID=cust_id, SKU_ID=sku_id, CustSKUCode=cust_sku))
            created += 1
    if not validate_only:
        model.commit()
    else:
        model.rollback()
    return {"ok": True, "master":"mappings", "created":created, "updated":updated, "skipped":skipped, "errors":errors}