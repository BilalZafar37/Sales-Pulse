# blueprints/soh.py
from flask import Blueprint, request, jsonify, send_file, current_app, render_template
from io import BytesIO
# Excel
from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting import Rule

from datetime import date, datetime
try:
    from zoneinfo import ZoneInfo
except ImportError:
    ZoneInfo = None
import pandas as pd
import hashlib, os
from collections import defaultdict
from sqlalchemy.exc import IntegrityError
from sqlalchemy import func, text

from models import model
from blueprints.sell_out_blueprint.sell_out import _resolve_sku_id  # reuse

from models import (
    SP_Customer, SP_SKU, SP_Customer_SKU_Map, Brands, SP_Status,
)
from models import SP_SOH_Uploads, SP_SOH_Detail, SP_InventoryLedger  # <-- only these for SOH flows
from config import STATIC_DIR, BASE_DIR, TZ_RIYADH

ALLOWED = {"xlsx","xls","csv"}


bp = Blueprint("soh", __name__, static_folder=STATIC_DIR, url_prefix="/soh")

def _begin_tx(session):
    return session.begin_nested() if session.in_transaction() else session.begin()

def _allowed(name:str)->bool:
    return "." in name and name.rsplit(".",1)[1].lower() in ALLOWED

def _sha256_fs(fs) -> str:
    fs.stream.seek(0)
    data = fs.read()          # IMPORTANT: read the FileStorage, not fs.stream
    fs.stream.seek(0)
    return hashlib.sha256(data).hexdigest()

def _normalize_soh_columns(df: pd.DataFrame) -> pd.DataFrame:
    rename_map, low = {}, {c.lower(): c for c in df.columns}
    for k in ["brand"]: 
        if k in low: rename_map[low[k]] = "Brand"; break
    for k in ["cust-sku","custsku","cust sku","custskucode"]:
        if k in low: rename_map[low[k]] = "CustSKUCode"; break
    for k in ["mec-sku","mec sku","articlecode","material","sku","mec"]:
        if k in low: rename_map[low[k]] = "MEC_SKU"; break
    for k in ["sohqty","qty","quantity","stockonhand"]:
        if k in low: rename_map[low[k]] = "SOHQty"; break
    for k in ["rownumber","row","#"]:
        if k in low: rename_map[low[k]] = "RowNumber"; break
    return df.rename(columns=rename_map)

def _load_df(fs)->pd.DataFrame:
    ext = fs.filename.rsplit(".",1)[1].lower()
    if ext in {"xlsx","xls"}:
        try:
            df = pd.read_excel(fs, sheet_name="SOH")
        except Exception:
            fs.stream.seek(0); df = pd.read_excel(fs)
    else:
        df = pd.read_csv(fs)
    df.columns = [c.strip() for c in df.columns]
    return _normalize_soh_columns(df)

def _article_exists_in_mCSI_exact(article_name: str) -> bool:
    if not article_name: 
        return False
    sql = text("""
        SELECT TOP 1 1
        FROM TB_WH_B2B_SO WITH (NOLOCK)
        WHERE LOWER(Article) = LOWER(:name)
    """)
    row = model.execute(sql, {"name": article_name.strip()}).first()
    return bool(row)

def _get_or_create_sku(brand: str, article: str) -> int | None:
    if not brand or not article:
        return None
    # Try exact (brand + article) match in SP_SKU
    sku = (model.query(SP_SKU)
           .filter(SP_SKU.Brand==brand.strip(),
                   SP_SKU.ArticleCode==article.strip()) 
           .first())
    if sku:
        return int(sku.SKU_ID)

    # Create if not exists
    sku = SP_SKU(
        Brand = brand.strip(),
        ArticleCode   = article.strip(),
    )
    model.add(sku)
    model.flush()
    return int(sku.SKU_ID)

def _ensure_customer_sku_map(customer_id: int, sku_id: int, cust_sku_code: str | None):
    """
    Ensure mapping exists only if a valid CustSKUCode is provided (not NaN, None, '', or 0).
    """
    # Normalize / check for invalid or missing codes
    if not cust_sku_code or str(cust_sku_code).strip() in ("", "0", "nan", "None"):
        # Skip creation silently
        return

    cust_sku_code = str(cust_sku_code).strip()

    exists = (model.query(SP_Customer_SKU_Map)
              .filter(SP_Customer_SKU_Map.CustomerID == customer_id,
                      SP_Customer_SKU_Map.SKU_ID == sku_id)
              .first())

    if exists:
        # Optionally update if code changed and existing is blank
        if not (exists.CustSKUCode or "").strip():
            exists.CustSKUCode = cust_sku_code
        return

    # Create only when valid code present
    model.add(SP_Customer_SKU_Map(
        CustomerID  = customer_id,
        SKU_ID      = sku_id,
        CustSKUCode = cust_sku_code,
        IsActive    = True,
        CreatedAt   = datetime.utcnow()
    ))

def _get_active_status_id():
    row = model.query(SP_Status).filter(SP_Status.StatusName == "Active").first()
    if row:
        return row.StatusID
    st = SP_Status(StatusName="Active")
    model.add(st); model.flush()
    return st.StatusID



@bp.route("/choices", methods=["GET"])
def choices():
    customers = [
        {"id": c.CustomerID, "code": c.CustCode, "name": c.CustName, "level": c.LevelType}
        for c in model.query(SP_Customer).filter(SP_Customer.LevelType == 'HO').order_by(SP_Customer.CustName).all()
    ]
    # brands = [{"name": b[0]} for b in model.query(Brands.BrandName).distinct().all() if b[0]]
    return jsonify({"customers": customers})

@bp.route("/template", methods=["GET"])
def template():
    customer_id = request.args.get("customer_id", type=int)

    cust_code = None
    if customer_id:
        cust = model.query(SP_Customer).filter(SP_Customer.CustomerID == customer_id).first()
        if cust:
            cust_code = cust.CustCode

    cust_part = cust_code if cust_code else f"CUST{customer_id or 'NA'}"
    today_str = datetime.now().strftime("%d-%b-%Y").upper()
    file_name = f"SOH_BY_BRAND_{cust_part}_{today_str}.xlsx"


    # 1) Base workbook via pandas
    df = pd.DataFrame(columns=["Brand", "Cust-SKU", "MEC-SKU", "SOHQty"])
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as xw:
        df.to_excel(xw, index=False, sheet_name="SOH")
    bio.seek(0)

    # 2) Load with openpyxl and add the lists/validation/formatting
    wb = load_workbook(bio)
    ws = wb["SOH"]

    # Fetch + sanitize brands
    raw = model.query(Brands).order_by(Brands.BrandName.asc()).all()
    brands = sorted({(b.BrandName or "").strip() for b in raw if (b.BrandName or "").strip()})

    if brands:
        # 2a) Put brands on a second sheet
        ws_list = wb.create_sheet("Lists")
        ws_list["A1"] = "BrandName"
        for r, name in enumerate(brands, start=2):
            ws_list.cell(row=r, column=1, value=name)

        # absolute range for validation list
        last_row = 1 + len(brands)
        list_range_abs = f"Lists!$A$2:$A${last_row}"

        # 2b) Data validation (dropdown) — direct range (no named range)
        dv = DataValidation(
            type="list",
            formula1=f"={list_range_abs}",
            allow_blank=True,               # allow blank rows; set False to force selection
            showErrorMessage=True,
            errorTitle="Invalid Brand",
            error="Please choose a Brand from the dropdown."
        )
        ws.add_data_validation(dv)
        dv.add("A2:A1048576")  # apply to entire Brand column

        # 2c) Spell-check-like highlight: anything not in list → light red fill
        dxf = DifferentialStyle(fill=PatternFill(fill_type="solid", start_color="FFF8D7DA", end_color="FFF8D7DA"))
        rule = Rule(type="expression", dxf=dxf, stopIfTrue=True)
        # COUNTIF against whole column on Lists; A2 is relative to each row in the applied range
        rule.formula = ["COUNTIF(Lists!$A:$A, A2)=0"]
        ws.conditional_formatting.add("A2:A1048576", rule)

        # Hide helper sheet
        ws_list.sheet_state = "hidden"

    # Nice UX
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 12

    # 3) Return the enriched workbook
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(
        out,  # IMPORTANT: return the enriched buffer
        as_attachment=True,
        download_name=file_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _get_prior_active_qty_for_day(customer_id:int, brand:str|None, sku_id:int, soh_date:date) -> float:
    """
    Read the prior 'active' snapshot qty for this (customer, brand?, sku, date).
    If multiple older headers existed (shouldn't), prefer the latest via upload id.
    """
    q = (model.query(SP_SOH_Detail.SOHQty)
         .join(SP_SOH_Uploads, SP_SOH_Uploads.SOHUploadID==SP_SOH_Detail.SOHUploadID)
         .filter(SP_SOH_Uploads.CustomerID==customer_id,
                 SP_SOH_Detail.SKU_ID==sku_id,
                 SP_SOH_Detail.SOHDate==soh_date,
                 SP_SOH_Detail.IsActive==True))
    if brand is not None:
        q = q.filter(SP_SOH_Uploads.Brand==brand)
    # choose latest header first
    q = q.order_by(SP_SOH_Uploads.SOHUploadID.desc()) 
    row = q.first()
    return float(row[0]) if row else 0.0

def _post_ledger_adjust(customer_id:int, sku_id:int, d:date, new_qty:float, old_qty:float,
                        hdr_id:int, row_num:int, movement_type:str):
    """
    Create a ledger snapshot delta row.
    movement_type:
      - "ADJUST"    for first-time / initial brand SOH
      - "SUPERCEED" for later corrections / re-uploads for the same brand
    """
    diff = float(new_qty) - float(old_qty)
    # You *can* skip zero-diff if you want to reduce noise
    idem = f"SOH_{movement_type}:{hdr_id}:{row_num}"
    model.add(SP_InventoryLedger(
        CustomerID=customer_id,
        SKU_ID=sku_id,
        DocDate=d,
        MovementType=movement_type,
        MovementSubType=None,
        Qty=diff,
        UploadID=None,
        RefTable="SP_SOH_Uploads",
        RefID=str(hdr_id),
        IdempotencyKey=idem,
        CreatedAt=datetime.now(TZ_RIYADH)
    ))


def _get_form_date_from_request() -> date | None:
    raw = (request.form.get("sohDate") or "").strip()
    if not raw:
        return None
    try:
        # parse with pandas, then cast to .date()
        return pd.to_datetime(raw, errors="raise").date()
    except Exception:
        # fallback manual parsing
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d-%b-%Y"):
            try:
                return datetime.strptime(raw, fmt).date()
            except Exception:
                pass
    return None

@bp.route("/upload", methods=["POST"])
def upload():
    try:
        customer_id   = request.form.get("customer_id", type=int)
        created_by    = (request.form.get("created_by") or "").strip() or None
        snapshot_type = (request.form.get("snapshot_type") or "Initial").strip()
        the_date      = _get_form_date_from_request()
        fs            = request.files.get("file")

        if not customer_id:
            return jsonify(ok=False, error="customer_id is required"), 400
        if not fs or not _allowed(fs.filename):
            return jsonify(ok=False, error="Upload a .xlsx/.xls/.csv file"), 400
        if not the_date:
            return jsonify(ok=False, error="Provide snapshot date in the form."), 400
        if the_date > date.today():
            return jsonify(ok=False, error="Future dates are not allowed for SOH."), 400

        
        
        
        file_hash = _sha256_fs(fs)
        
        # Debug: see any row with this hash in the table (across all customers)
        # try:
        #     rows = model.execute(
        #         text("""
        #             SELECT TOP 5 SOHUploadID, CustomerID, [Date], SourceFileHash
        #             FROM dbo.SP_SOH_Uploads WITH (NOLOCK)
        #             WHERE SourceFileHash = :h
        #             ORDER BY SOHUploadID DESC
        #         """),
        #         {"h": file_hash}
        #     ).fetchall()
        #     current_app.logger.info("DEBUG HASH %s -> existing rows: %s", file_hash, rows)
        # except Exception as _e:
        #     current_app.logger.exception("DEBUG HASH lookup failed")
        
        dup = (model.query(SP_SOH_Uploads.SOHUploadID)
               .filter(SP_SOH_Uploads.CustomerID==customer_id,
                       SP_SOH_Uploads.SourceFileHash==file_hash)
               .first())
        if dup:
            return jsonify(ok=False, error="This exact file was already uploaded for this customer."), 400

        df = _load_df(fs).copy()
        df.columns = [c.strip() for c in df.columns]

        required_cols = {"Brand","MEC_SKU","SOHQty"}
        missing = [c for c in required_cols if c not in df.columns]
        if missing:
            return jsonify(ok=False, error=f"Missing required columns: {', '.join(missing)}"), 400
        
        # ---- Validate and normalize Brand names ----
        # Load master list of valid brand names from DB
        valid_brands = { (b.BrandName or "").strip().upper() for b in model.query(Brands).all() if (b.BrandName or "").strip() }
        valid_lower_map = { vb.lower(): vb for vb in valid_brands }  # lowercase -> canonical uppercase
        row_errors = []
        
        for _, r in df.iterrows():
            rn = int(r.get("RowNumber") or 0)
            brand = (str(r.get("Brand") or "").strip())
            if not brand:
                row_errors.append(f"Row {rn}: Brand is required.")
                continue
        
            brand_lower = brand.lower()
            if brand_lower not in valid_lower_map:
                # spelling or unrecognized brand error
                row_errors.append(
                    f'Row {rn}: Invalid brand/Spelling  "{brand}". '
                    # f'Allowed brands are: {", ".join(sorted(valid_brands))}.'
                )
            else:
                # fix to uppercase canonical
                df.at[_, "Brand"] = valid_lower_map[brand_lower]
        
        if row_errors:
            return jsonify(ok=False, error="; ".join(row_errors)), 400
        
        
                # ---- Collect all brands present in this file ----
        upload_brands = sorted({
            (str(b).strip())
            for b in df["Brand"].dropna().unique()
            if str(b).strip()
        })
        if not upload_brands:
            return jsonify(ok=False, error="No valid Brand found in the file."), 400

        # For 'Initial' uploads: block if any of these brands already has a non-superseded Initial
        if snapshot_type.lower() == "initial":
            exists_initial = (
                model.query(SP_SOH_Uploads.SOHUploadID)
                .filter(
                    SP_SOH_Uploads.CustomerID == customer_id,
                    func.lower(SP_SOH_Uploads.SnapshotType) == "initial",
                    SP_SOH_Uploads.Brand.in_(upload_brands),
                    SP_SOH_Uploads.Status != "Superseded"
                )
                .first()
            )
            if exists_initial:
                return jsonify(
                    ok=False,
                    error=(
                        "Initial SOH already exists for this customer for at least one of these brands: "
                        + ", ".join(upload_brands)
                        + ". Please use 'Supersede' snapshot type instead."
                    )
                ), 400
        

        df["SOHQty"] = pd.to_numeric(df["SOHQty"], errors="coerce")
        if "RowNumber" not in df.columns:
            df["RowNumber"] = range(1, len(df)+1)
        df = df.dropna(subset=["SOHQty"])
        if df.empty:
            return jsonify(ok=False, error="No valid rows (SOHQty missing)"), 400
        
        # ----- Block negative SOHQty with row numbers -----
        neg_rows = df.loc[pd.to_numeric(df["SOHQty"], errors="coerce") < 0, ["RowNumber", "MEC_SKU", "SOHQty"]]
        if not neg_rows.empty:
            msgs = [f'Row {int(r.RowNumber)}: SOHQty cannot be negative (value {r.SOHQty}). Article: {str(r.MEC_SKU).strip()}'
                    for _, r in neg_rows.iterrows()]
            return jsonify(ok=False, error="; ".join(msgs)), 400
        
        bad = df.loc[(df["SOHQty"] % 1) != 0, ["RowNumber","MEC_SKU","SOHQty"]]
        if not bad.empty:
            msgs = [f'Row {int(x.RowNumber)}: QTY must be a whole number (got {x.SOHQty}). Article: {x.MEC_SKU}'
                    for _, x in bad.iterrows()]
            return jsonify(ok=False, error="; ".join(msgs)), 400

        # --- Validate Articles (exact match in RTOS_MCSI since 2023) ---
        row_errors = []
        for _, r in df.iterrows():
            rn      = int(r.get("RowNumber") or 0)
            brand   = (r.get("Brand") or "").strip()
            article = str(r.get("MEC_SKU") or "").strip()

            if not brand:
                row_errors.append(f'Row {rn}: Brand is required.')
                continue
            if not article:
                row_errors.append(f'Row {rn}: Article is required.')
                continue
            if not _article_exists_in_mCSI_exact(article):
                row_errors.append(
                    f'Article on excel row {rn} named "{article}" does not exist in MCSI since 2023 or there is a spelling mismatch'
                )

        if row_errors:
            return jsonify(ok=False, error="; ".join(row_errors)), 400

        # --- Resolve/Create SKUs & group detail records by Brand ---
        # from collections import defaultdict
        records_by_brand: dict[str, list[dict]] = defaultdict(list)

        with _begin_tx(model):
            # 1) Build records per brand
            for _, r in df.iterrows():
                rn        = int(r["RowNumber"])
                brand     = (r.get("Brand") or "").strip()
                article   = (r.get("MEC_SKU") or "").strip()
                cust_code = str(r.get("CustSKUCode") or r.get("Cust-SKU") or "").strip() or None
                qty       = float(r.get("SOHQty") or 0.0)

                if not brand:
                    row_errors.append(f"Row {rn}: Brand is required.")
                    continue

                sku_id = _get_or_create_sku(brand, article)
                if not sku_id:
                    row_errors.append(f"Row {rn}: Unable to create/resolve SKU for [{brand}] / [{article}].")
                    continue

                _ensure_customer_sku_map(customer_id, sku_id, cust_code)

                records_by_brand[brand].append({
                    "SKU_ID":    int(sku_id),
                    "SOHQty":    qty,
                    "RowNumber": rn,
                })

            if row_errors:
                model.rollback()
                return jsonify(ok=False, error="; ".join(row_errors)), 400

            deactivated_total = 0
            inserted_total    = 0
            superseded_ids_total: list[int] = []
            first_hdr_id = None

            # Decide ledger movement type for this upload
            # - "initial" => ADJUST
            # - other (e.g. Supersede) => SUPERCEED
            movement_type = "ADJUST" if snapshot_type.lower() == "initial" else "SUPERCEED"

            # 2) Process each brand chunk separately: one header per (customer, brand, date)
            for brand in sorted(records_by_brand.keys()):
                brand_records = records_by_brand[brand]
                if not brand_records:
                    continue

                # 2a) Deactivate prior active details for this customer+brand+date
                prior_hdr_q = (
                    model.query(SP_SOH_Uploads)
                    .filter(
                        SP_SOH_Uploads.CustomerID == customer_id,
                        SP_SOH_Uploads.Date == the_date,
                        SP_SOH_Uploads.Brand == brand,
                    )
                )
                prior_headers = prior_hdr_q.all()
                prior_hdr_ids = [h.SOHUploadID for h in prior_headers]

                if prior_hdr_ids:
                    deactivated = (
                        model.query(SP_SOH_Detail)
                        .filter(
                            SP_SOH_Detail.SOHUploadID.in_(prior_hdr_ids),
                            SP_SOH_Detail.SOHDate == the_date,
                            SP_SOH_Detail.IsActive == True,
                        )
                        .update({SP_SOH_Detail.IsActive: False}, synchronize_session=False)
                    )
                    deactivated_total += int(deactivated)

                # 2b) New header for this brand
                hdr = SP_SOH_Uploads(
                    CustomerID     = customer_id,
                    SnapshotType   = snapshot_type,
                    Brand          = brand,
                    Date           = the_date,
                    Status         = "Draft",
                    CreatedBy      = created_by,
                    CreatedAt      = datetime.now(tz=ZoneInfo("Asia/Riyadh") if ZoneInfo else None),
                    SourceFileName = fs.filename,
                    SourceFileHash = file_hash,
                )
                model.add(hdr)
                model.flush()

                if first_hdr_id is None:
                    first_hdr_id = hdr.SOHUploadID

                for ph in prior_headers:
                    ph.Status = "Superseded"
                    ph.SupersededByUploadID = hdr.SOHUploadID
                    superseded_ids_total.append(ph.SOHUploadID)

                # 2c) Details + ledger for this brand
                detail_objs = []
                for rec in brand_records:
                    old_qty = _get_prior_active_qty_for_day(customer_id, brand, rec["SKU_ID"], the_date)

                    detail_objs.append(SP_SOH_Detail(
                        SOHUploadID = hdr.SOHUploadID,
                        SKU_ID      = rec["SKU_ID"],
                        RowNumber   = rec["RowNumber"],
                        SOHDate     = the_date,
                        SOHQty      = rec["SOHQty"],
                        IsActive    = True,
                    ))

                    _post_ledger_adjust(
                        customer_id = customer_id,
                        sku_id      = rec["SKU_ID"],
                        d           = the_date,
                        new_qty     = rec["SOHQty"],
                        old_qty     = old_qty,
                        hdr_id      = hdr.SOHUploadID,
                        row_num     = rec["RowNumber"],
                        movement_type = movement_type,
                    )

                model.add_all(detail_objs)
                inserted_total += len(detail_objs)

            # 3) Mark customer Active if it has at least one ADJUST row
            has_adjust = (
                model.query(SP_InventoryLedger.CustomerID)
                .filter(
                    SP_InventoryLedger.CustomerID == customer_id,
                    SP_InventoryLedger.MovementType == "ADJUST",
                )
                .first()
            )
            if has_adjust:
                active_id = _get_active_status_id()
                model.query(SP_Customer).filter(
                    SP_Customer.CustomerID == customer_id,
                    (SP_Customer.StatusID != active_id) | (SP_Customer.StatusID.is_(None)),
                ).update(
                    {
                        SP_Customer.StatusID: active_id,
                        SP_Customer.StatusDate: the_date,
                    },
                    synchronize_session=False,
                )

            

        model.commit()
        return jsonify(
            ok=True,
            soh_upload_id  = first_hdr_id,          # primary header id (first brand)
            soh_upload_ids = superseded_ids_total and [first_hdr_id] + superseded_ids_total or [first_hdr_id],
            snapshot_type  = snapshot_type,
            date           = str(the_date),
            inserted       = int(inserted_total),
            deactivated    = int(deactivated_total),
            superseded     = superseded_ids_total,
        )


    except IntegrityError as e:
        model.rollback()
        current_app.logger.exception("IntegrityError on SOH upload")
        return jsonify(ok=False, error=f"Integrity error: {getattr(e.orig, 'args', [str(e)])[:1][0]}"), 400
    except Exception as e:
        model.rollback()
        current_app.logger.exception("SOH upload failed")
        return jsonify(ok=False, error=str(e)), 500


@bp.route("/", methods=["GET", "POST"])
def front():
    return render_template('./soh/soh.html')
