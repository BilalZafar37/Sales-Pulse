# blueprints/customer_management.py
from __future__ import annotations
from flask import Blueprint, request, jsonify, send_file, current_app, render_template
from datetime import datetime, date, timedelta
from sqlalchemy import or_, func, text
import io
from openpyxl import Workbook, load_workbook

from models import model, SP_Customer, SP_Status, SP_InventoryLedger, SP_GlobalConfig, SP_CustomerStatusTag
from config import STATIC_DIR

bp = Blueprint(
    "customer_management_bp",__name__, url_prefix="/customer-mgmt", static_folder = STATIC_DIR
)

# ----------------------------
# Helpers
# ----------------------------

def _status_id_from_payload(data):
    sid = data.get("StatusID")
    if sid:
        return int(sid)
    sname = (data.get("Status") or data.get("StatusName") or "").strip()
    if not sname:
        return None
    row = model.query(SP_Status.StatusID).filter(SP_Status.StatusName == sname).first()
    if not row:
        # auto-create if you prefer; else return None to reject
        st = SP_Status(StatusName=sname)
        model.add(st); model.commit()
        return st.StatusID
    return row[0]

def _parse_int(v, default=None):
    try:
        return int(v)
    except Exception:
        return default

def _paginate(q, default_page=1, default_size=50):
    page = _parse_int(request.args.get("page"), default_page) or default_page
    page_size = _parse_int(request.args.get("page_size"), default_size) or default_size
    total = q.count()
    rows = q.offset((page-1)*page_size).limit(page_size).all()
    return total, rows

def _today():
    return date.today()

def _json_date(d):
    if not d:
        return None
    if isinstance(d, (datetime, )):
        return d.date().isoformat()
    if isinstance(d, date):
        return d.isoformat()
    return str(d)

# ============
def _get_cfg_int(key: str, default: int) -> int:
    row = model.query(SP_GlobalConfig.Value).filter(SP_GlobalConfig.Key == key).first()
    if not row: return default
    try: return int(row[0])
    except Exception: return default

def _get_status_id(name: str) -> int:
    s = model.query(SP_Status).filter(SP_Status.StatusName == name).first()
    if s: return s.StatusID
    s = SP_Status(StatusName=name)
    model.add(s); model.commit()
    return s.StatusID

def _sync_tags(customer_id: int, want_ids: set[int]):
    """
    Make SP_CustomerStatusTag exactly match want_ids for this customer,
    using SQL Server MERGE for insert-if-missing (atomic) and a delete pass
    for stale tags. Safe under concurrency; will not raise PK violations.
    """
    # Normalize to sorted unique ints
    want = sorted(int(s) for s in want_ids if s)

    # 1) Delete anything not desired
    if want:
        in_list = ",".join(str(s) for s in want)
        model.execute(text(f"""
            DELETE T
            FROM SP_CustomerStatusTag AS T
            WHERE T.CustomerID = :cid
              AND T.StatusID NOT IN ({in_list})
        """), {"cid": customer_id})
    else:
        # If want is empty: remove all tags for this customer
        model.execute(text("""
            DELETE FROM SP_CustomerStatusTag WHERE CustomerID = :cid
        """), {"cid": customer_id})

    # 2) Insert-if-missing using MERGE for each desired tag
    for sid in want:
        model.execute(text("""
            MERGE INTO SP_CustomerStatusTag AS tgt
            USING (SELECT :cid AS CustomerID, :sid AS StatusID) AS src
              ON (tgt.CustomerID = src.CustomerID AND tgt.StatusID = src.StatusID)
            WHEN NOT MATCHED THEN
              INSERT (CustomerID, StatusID) VALUES (src.CustomerID, src.StatusID);
        """), {"cid": customer_id, "sid": sid})


def _recompute_customer_statuses() -> dict:
    today = date.today()
    dead_days   = _get_cfg_int("DeadThresholdDays",               90)
    hib_in_days = _get_cfg_int("HibernatingSellInThresholdDays",  30)
    hib_out_days= _get_cfg_int("HibernatingSellOutThresholdDays", 30)

    sid_active   = _get_status_id("Active")
    sid_dead     = _get_status_id("DEAD")
    sid_disabled = _get_status_id("Disabled")
    tag_in       = _get_status_id("Hibernating-Sell-in")
    tag_out      = _get_status_id("Hibernating-Sell-out")

    # Subqueries: last SELLIN/SELLOUT
    last_in = (model.query(
                    SP_InventoryLedger.CustomerID.label("CID"),
                    func.max(SP_InventoryLedger.DocDate).label("last_in"))
               .filter(SP_InventoryLedger.MovementType == "SELLIN")
               .group_by(SP_InventoryLedger.CustomerID)
               ).subquery()

    last_out = (model.query(
                    SP_InventoryLedger.CustomerID.label("CID"),
                    func.max(SP_InventoryLedger.DocDate).label("last_out"))
                .filter(SP_InventoryLedger.MovementType == "SELLOUT")
                .group_by(SP_InventoryLedger.CustomerID)
                ).subquery()

    rows = (model.query(
                SP_Customer.CustomerID,
                SP_Customer.StatusID,
                last_in.c.last_in,
                last_out.c.last_out)
            .outerjoin(last_in, last_in.c.CID == SP_Customer.CustomerID)
            .outerjoin(last_out, last_out.c.CID == SP_Customer.CustomerID)
            ).all()

    tally = {"PrimaryUpdated":0, "TagsUpdated":0, "SkippedDisabled":0}
    for cid, cur_sid, d_in, d_out in rows:
        # Skip auto for Disabled
        if cur_sid == sid_disabled:
            tally["SkippedDisabled"] += 1
            continue

        ds_in  = (today - d_in).days  if d_in  else 10**9
        ds_out = (today - d_out).days if d_out else 10**9

        # Compute tags independently (can hold both)
        want_tags = set()
        if ds_in  > hib_in_days:  want_tags.add(tag_in)
        if ds_out > hib_out_days: want_tags.add(tag_out)

        # DEAD only when BOTH are past Dead threshold; else ACTIVE
        new_primary = sid_dead if (ds_in > dead_days and ds_out > dead_days) else sid_active

        # Upsert tags
        before = set(
            sid for (sid,) in model.query(SP_CustomerStatusTag.StatusID)
                                   .filter(SP_CustomerStatusTag.CustomerID == cid)
        )
        if before != want_tags:
            # Compute tags independently (can hold both)
            want_tags = set()
            if ds_in  > hib_in_days:  want_tags.add(tag_in)
            if ds_out > hib_out_days: want_tags.add(tag_out)
    
            # Sync tags (atomic + idempotent)
            _sync_tags(cid, want_tags)
            tally["TagsUpdated"] += 1  # optional: count only when changed; fine to leave as-is

        # Update primary if changed
        if cur_sid != new_primary:
            model.query(SP_Customer)\
                 .filter(SP_Customer.CustomerID == cid)\
                 .update({
                     SP_Customer.StatusID: new_primary,
                     SP_Customer.StatusDate: today
                 }, synchronize_session=False)
            tally["PrimaryUpdated"] += 1

    model.commit()
    return {
        "ok": True,
        "tally": tally,
        "thresholds": {
            "Dead": dead_days,
            "HibernatingSellIn": hib_in_days,
            "HibernatingSellOut": hib_out_days
        }
    }


@bp.route("/api/status/recompute", methods=["POST"])
def status_recompute():
    res = _recompute_customer_statuses()
    return jsonify(ok=True, **res)


# ----------------------------
# Options: customers (for Select2)
# ----------------------------

@bp.route("/api/options/customers", methods=["GET"])
def options_customers():
    term = (request.args.get("term") or "").strip()
    level = (request.args.get("level") or "").strip().upper()  # e.g., HO only
    q = model.query(SP_Customer)
    if term:
        like = f"%{term}%"
        q = q.filter(or_(SP_Customer.CustName.ilike(like), SP_Customer.CustCode.ilike(like)))
    if level in ("HO", "BRANCH"):
        q = q.filter(SP_Customer.LevelType == level)
    q = q.order_by(SP_Customer.CustName).limit(50)
    items = [{"id": r.CustomerID, "text": f"{r.CustCode} — {r.CustName} ({r.LevelType})"} for r in q]
    return jsonify({"results": items})

# ----------------------------
# Customers list (with parent, status, status date)
# ----------------------------

@bp.route("/api/customers", methods=["GET"])
def customers_list():
    term = (request.args.get("q") or "").strip()

    q = (model.query(
            SP_Customer.CustomerID,
            SP_Customer.CustCode,
            SP_Customer.CustName,
            SP_Customer.LevelType,
            SP_Customer.ParentCustID,
            SP_Customer.StatusID,
            SP_Status.StatusName,          # <- bring as column
            SP_Customer.StatusDate
        )
        .outerjoin(SP_Status, SP_Status.StatusID == SP_Customer.StatusID)
    )

    if term:
        like = f"%{term}%"
        q = q.filter(or_(SP_Customer.CustName.ilike(like), SP_Customer.CustCode.ilike(like)))

    q = q.order_by(SP_Customer.CustName)
    total, rows = _paginate(q)

    parent_ids = {r.ParentCustID for r in rows if r.ParentCustID}
    parent_map = {}
    if parent_ids:
        pairs = (model.query(SP_Customer.CustomerID, SP_Customer.CustCode, SP_Customer.CustName)
                      .filter(SP_Customer.CustomerID.in_(parent_ids)).all())
        parent_map = {cid: (code, name) for cid, code, name in pairs}

    page_ids = [r.CustomerID for r in rows]
    tag_rows = (model.query(SP_CustomerStatusTag.CustomerID, SP_Status.StatusName)
                .join(SP_Status, SP_Status.StatusID == SP_CustomerStatusTag.StatusID)
                .filter(SP_CustomerStatusTag.CustomerID.in_(page_ids))
                .all())
    tags_by_cust = {}
    for cid, sname in tag_rows:
        tags_by_cust.setdefault(cid, []).append(sname)
    
    items = []
    for r in rows:
        # r is a tuple-like row; attribute access works via labels
        p_code, p_name = (None, None)
        if r.ParentCustID and r.ParentCustID in parent_map:
            p_code, p_name = parent_map[r.ParentCustID]

        items.append({
            "CustomerID":   r.CustomerID,
            "CustCode":     r.CustCode,
            "CustName":     r.CustName,
            "LevelType":    r.LevelType,
            "ParentCustID": r.ParentCustID,
            "ParentCustCode": p_code,
            "ParentCustName": p_name,
            "StatusID":     r.StatusID,
            "StatusName":   r.StatusName,                    # <- explicit
            "StatusDate":   _json_date(r.StatusDate),
            "StatusTags": tags_by_cust.get(r.CustomerID, []),
        })
    return jsonify(ok=True, total=total, items=items)


# ----------------------------
# Create / Update / Delete
# ----------------------------

@bp.route("/api/customers", methods=["POST"])
def customers_create():
    data = request.get_json(force=True)
    c = SP_Customer(
        CustCode=data["CustCode"].strip(),
        CustName=data["CustName"].strip(),
        LevelType=(data.get("LevelType") or "HO").strip(),
        ParentCustID=data.get("ParentCustID"),
    )
    c.StatusID = _status_id_from_payload(data) or c.StatusID  # falls back to DB default
    if "StatusDate" in data:
        try:
            c.StatusDate = date.fromisoformat(data["StatusDate"]) if data["StatusDate"] else None
        except Exception:
            c.StatusDate = None

    model.add(c)
    model.commit()
    return jsonify(ok=True, id=c.CustomerID)

@bp.route("/api/customers/<int:cid>", methods=["PATCH"])
def customers_update(cid):
    data = request.get_json(force=True)
    c = model.query(SP_Customer).get(cid)
    if not c:
        return jsonify(ok=False, error="Not found"), 404

    for k in ("CustCode","CustName","LevelType","ParentCustID"):
        if k in data:
            setattr(c, k, data[k])

    sid = _status_id_from_payload(data)
    if sid is not None:
        c.StatusID = sid
        if "StatusDate" not in data:   # stamp today if caller changed status but gave no date
            c.StatusDate = _today()
    
    if "StatusDate" in data:
        v = data.get("StatusDate")
        try:
            c.StatusDate = (date.fromisoformat(v) if v else None)
        except Exception:
            pass
    
    model.commit()
    return jsonify(ok=True)

@bp.route("/api/customers/<int:cid>", methods=["DELETE"])
def customers_delete(cid):
    c = model.query(SP_Customer).get(cid)
    if not c:
        return jsonify(ok=False, error="Not found"), 404
    model.delete(c)
    model.commit()
    return jsonify(ok=True)

# ----------------------------
# Parent (HO) assignment helpers
# ----------------------------

def _is_ho(customer_id: int) -> bool:
    r = model.query(SP_Customer.LevelType).filter(SP_Customer.CustomerID == customer_id).first()
    return bool(r and r[0] == "HO")

@bp.route("/api/customers/<int:child_id>/set_parent", methods=["POST"])
def set_parent(child_id):
    """
    Set or clear parent for a single child.
    payload: {"parent_id": <int or null>}
    Rules:
      - parent must be HO (if provided)
      - cannot set parent to self
    """
    data = request.get_json(force=True)
    parent_id = data.get("parent_id")
    child = model.query(SP_Customer).get(child_id)
    if not child:
        return jsonify(ok=False, error="Child not found"), 404

    if parent_id is None:
        child.ParentCustID = None
        model.commit()
        return jsonify(ok=True)

    if parent_id == child_id:
        return jsonify(ok=False, error="Cannot parent to self"), 400

    if not _is_ho(parent_id):
        return jsonify(ok=False, error="Parent must be HO"), 400

    child.ParentCustID = parent_id
    model.commit()
    return jsonify(ok=True)

@bp.route("/api/customers/reparent_children", methods=["POST"])
def reparent_children():
    """
    Bulk reparent: move all children of one HO to another HO.
    payload: {"old_parent_id": X, "new_parent_id": Y}
    Both must be HO; allows merging branches under a new HO.
    """
    data = request.get_json(force=True)
    old_pid = data.get("old_parent_id")
    new_pid = data.get("new_parent_id")

    if not (old_pid and new_pid):
        return jsonify(ok=False, error="old_parent_id and new_parent_id are required"), 400
    if old_pid == new_pid:
        return jsonify(ok=False, error="old and new cannot be same"), 400
    if not _is_ho(old_pid) or not _is_ho(new_pid):
        return jsonify(ok=False, error="Both must be HO"), 400

    updated = (model.query(SP_Customer)
                     .filter(SP_Customer.ParentCustID == old_pid)
                     .update({SP_Customer.ParentCustID: new_pid}, synchronize_session=False))
    model.commit()
    return jsonify(ok=True, moved=updated)

# ----------------------------
# Bulk Excel: customers only
# ----------------------------

MASTER_SPEC = {
    "sheet": "Customers",
    "columns": [
        ("CustCode",   True,  "Unique customer code"),
        ("CustName",   True,  "Customer name"),
        ("LevelType",  True,  "HO or Branch"),
        ("ParentCode", False, "Parent customer's CustCode (if Branch)"),
        ("Status",     False, "e.g., Active, Inactive, On Hold"),
        ("StatusDate", False, "YYYY-MM-DD"),
    ],
    "sample": [
        {"CustCode": "CUST1001", "CustName": "ABC Trading",          "LevelType": "HO",     "ParentCode": "",         "Status":"Active", "StatusDate": "2025-01-01"},
        {"CustCode": "CUST1002", "CustName": "ABC Retail Riyadh",    "LevelType": "Branch", "ParentCode": "CUST1001", "Status":"Active", "StatusDate": "2025-01-01"},
        {"CustCode": "CUST1003", "CustName": "ABC Retail Jeddah",    "LevelType": "Branch", "ParentCode": "CUST1001", "Status":"On Hold","StatusDate": "2025-03-15"},
    ],
}

def _make_template_wb() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = MASTER_SPEC["sheet"]
    # headers
    headers = [f"{c}{' *' if req else ''}" for c, req, _ in MASTER_SPEC["columns"]]
    ws.append(headers)
    # notes
    ws.append([note for _, _, note in MASTER_SPEC["columns"]])
    # sample
    for samp in MASTER_SPEC["sample"]:
        row = [samp.get(col, "") for col, *_ in MASTER_SPEC["columns"]]
        ws.append(row)
    # widths
    for i, (col, *_rest) in enumerate(MASTER_SPEC["columns"], start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(14, len(col) + 2)
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return bio.getvalue()

@bp.route("/api/bulk/template/customers", methods=["GET"])
def bulk_template_customers():
    content = _make_template_wb()
    return send_file(io.BytesIO(content), as_attachment=True,
                     download_name="Customers_Template.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def _ws_to_dicts(ws):
    col_names = [c for c, *_ in MASTER_SPEC["columns"]]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i <= 2:  # header + notes
            continue
        if all((v is None or str(v).strip() == "") for v in row):
            continue
        item = {}
        for j, name in enumerate(col_names):
            v = (row[j] if j < len(row) else None)
            item[name] = v.strip() if isinstance(v, str) else v
        rows.append(item)
    return rows

@bp.route("/api/bulk/upload/customers", methods=["POST"])
def bulk_upload_customers():
    f = request.files.get("file")
    if not f:
        return jsonify(ok=False, error="No file"), 400
    try:
        wb = load_workbook(filename=io.BytesIO(f.read()), data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify(ok=False, error=f"Invalid Excel: {e}"), 400

    rows = _ws_to_dicts(ws)
    required = {"CustCode","CustName","LevelType"}
    created = updated = skipped = 0
    errors = []

    existing = {c.CustCode.upper(): c for c in model.query(SP_Customer).all()}

    def _to_date(s):
        if not s: return None
        try:
            return date.fromisoformat(str(s))
        except Exception:
            return None

    for i, r in enumerate(rows, start=1):
        if any(not r.get(k) for k in required):
            skipped += 1
            errors.append({"row": i, "error": "Missing CustCode/CustName/LevelType"})
            continue
        code = (r["CustCode"] or "").strip().upper()
        name = (r["CustName"] or "").strip()
        level= (r["LevelType"] or "HO").strip()
        parent_code = (r.get("ParentCode") or "").strip().upper()
        status_name = (r.get("Status") or r.get("StatusName") or "").strip()
        status_id = None
        if status_name:
            status_id = _status_id_from_payload({"StatusName": status_name})
        sdate  = _to_date(r.get("StatusDate"))

        parent_id = None
        if parent_code:
            p = existing.get(parent_code)
            if not p:
                skipped += 1
                errors.append({"row": i, "error": f"Parent code not found: {parent_code}"})
                continue
            if p.LevelType != "HO":
                skipped += 1
                errors.append({"row": i, "error": f"Parent must be HO: {parent_code}"})
                continue
            parent_id = p.CustomerID

        obj = existing.get(code)
        if obj:
            obj.CustName     = name
            obj.LevelType    = level
            obj.ParentCustID = parent_id
            if status_id is not None:
                obj.StatusID = status_id
            if hasattr(obj, "StatusDate"):
                obj.StatusDate = sdate
            updated += 1
        else:
            obj = SP_Customer(
                CustCode=code, CustName=name, LevelType=level,
                ParentCustID=parent_id,
                StatusID=status_id  # <- just use resolved id (or None)
            )
            if hasattr(obj, "StatusDate"):
                obj.StatusDate = sdate
            model.add(obj); model.commit()
            existing[code] = obj
            created += 1
        
        

    model.commit()
    return jsonify(ok=True, created=created, updated=updated, skipped=skipped, errors=errors)

# ----------------------------
# Page route (separate frontend)
# ----------------------------

@bp.route("/", methods=["GET"])
def page():
    return render_template("customers.html")


@bp.route("/api/options/statuses", methods=["GET"])
def options_statuses():
    term = (request.args.get("term") or "").strip()
    q = model.query(SP_Status)
    if term:
        like = f"%{term}%"
        q = q.filter(SP_Status.StatusName.ilike(like))
    q = q.order_by(SP_Status.StatusName).limit(50)
    items = [{"id": r.StatusID, "text": r.StatusName} for r in q]
    return jsonify({"results": items})
