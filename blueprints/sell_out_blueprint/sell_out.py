from flask import Blueprint, request, jsonify, send_file, current_app, render_template
from collections import defaultdict
from io import BytesIO
from datetime import datetime, date, timedelta
from dateutil import parser as date_parser
import pandas as pd
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import hashlib
import os
from sqlalchemy.exc import IntegrityError

from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill


from models import model 
from models import (
    SP_Customer, SP_SKU, SP_Customer_SKU_Map,
    SP_SellOutUploads, SP_MCSI_SellOut,
    SP_SellOutUploadAudit, Brands, SP_MCSI_SellIn , SP_SOH_Detail, SP_SOH_Uploads
)
from config import STATIC_DIR, TZ_RIYADH

bp = Blueprint("sell_out", __name__, static_folder=STATIC_DIR, url_prefix="/sell-out")

ALLOWED = {"xlsx","xls","csv"}

# helper to open a transaction or savepoint safely
def _begin_tx(session):
    return session.begin_nested() if session.in_transaction() else session.begin()

def _allowed(name:str)->bool:
    return "." in name and name.rsplit(".",1)[1].lower() in ALLOWED

def _sha256_fs(fs)->str:
    pos = fs.stream.tell()
    fs.stream.seek(0)
    h = hashlib.sha256(fs.stream.read()).hexdigest()
    fs.stream.seek(pos)
    return h

def _load_df(fs)->pd.DataFrame:
    ext = fs.filename.rsplit(".",1)[1].lower()
    if ext in {"xlsx","xls"}:
        try:
            df = pd.read_excel(fs, sheet_name="SellOut")
        except Exception:
            fs.stream.seek(0)
            df = pd.read_excel(fs)
    else:
        df = pd.read_csv(fs)

    df.columns = [c.strip() for c in df.columns]
    df = _normalize_sellout_columns(df)

    # Types & coercion
    if "DocumentDate" in df.columns:
        # user sample is often dd-MM-yyyy; dayfirst=True is safer
        df["DocumentDate"] = pd.to_datetime(df["DocumentDate"], errors="coerce", dayfirst=True).dt.date
    if "SellOutQty" in df.columns:
        df["SellOutQty"] = pd.to_numeric(df["SellOutQty"], errors="coerce")
    if "ReportedSOH" in df.columns:
        df["ReportedSOH"] = pd.to_numeric(df["ReportedSOH"], errors="coerce")

    # Ensure RowNumber exists (backend-only)
    if "RowNumber" not in df.columns:
        df["RowNumber"] = range(1, len(df)+1)

    return df

def _to_date(x):
    if pd.isna(x) or x is None or x=="":
        return None
    return pd.to_datetime(x, errors="coerce").date()

def _resolve_sku_id(row, customer_id: int, cache: dict, *, create_mapping: bool = True):
    """
    Resolve SKU_ID from a row. Also enforce/maintain the (customer, MEC-SKU) <-> CustSKUCode mapping.

    Rules:
      - If BOTH ArticleCode (or MEC-SKU) AND CustSKUCode are provided:
          • If a mapping exists for (customer_id, sku_id), the code must match; else raise ValueError.
          • If no mapping exists, create it (unless the same CustSKUCode is mapped to another SKU -> error).
      - If ONLY ArticleCode is provided: resolve via SP_SKU.
      - If ONLY CustSKUCode is provided: resolve via SP_Customer_SKU_Map.
      - If SKU_ID is provided, use it; still validate/add mapping if CustSKUCode is provided.

    Returns: int SKU_ID or None (if unresolved and no error condition).
    Raises: ValueError on mapping conflicts or unknown MEC-SKU when a cust code is also present.
    """
    # Helpers to pull normalized values
    def _get(row, *keys):
        for k in keys:
            if k in row and pd.notna(row[k]):
                return str(row[k]).strip()
        return None

    cust_code = _get(row, "CustSKUCode", "Cust-SKU")
    art_val   = _get(row, "ArticleCode", "MEC-SKU")

    # 0) (Optional) explicit SKU_ID
    sku_id = None
    if "SKU_ID" in row and pd.notna(row["SKU_ID"]):
        try:
            sku_id = int(row["SKU_ID"])
        except Exception:
            sku_id = None

    # 1) Resolve by MEC article if needed
    if sku_id is None and art_val:
        key_art = ("art", art_val)
        if key_art not in cache:
            sku = model.query(SP_SKU.SKU_ID).filter(SP_SKU.ArticleCode == art_val).scalar()
            cache[key_art] = sku
        sku_id = cache[key_art]

    # 2) If we still don't have sku_id and we have a cust code → try mapping
    if sku_id is None and cust_code:
        key_cust = ("cust", customer_id, cust_code)
        if key_cust not in cache:
            mapped = (
                model.query(SP_Customer_SKU_Map.SKU_ID)
                .filter(SP_Customer_SKU_Map.CustomerID == customer_id,
                        SP_Customer_SKU_Map.CustSKUCode == cust_code)
                .scalar()
            )
            cache[key_cust] = mapped
        sku_id = cache[key_cust]

    # 3) If BOTH article and cust code are present, enforce mapping consistency
    if art_val and cust_code and cust_code not in ("", "0", "nan", "None"):
        # We must have a valid sku for the MEC article; otherwise tell the user
        if sku_id is None:
            raise ValueError(f"Unknown MEC-SKU '{art_val}'. Please ensure the SKU exists in SP_SKU before uploading.")

        # Check existing mapping for (customer, sku)
        key_pair = ("map_by_pair", customer_id, sku_id)
        if key_pair not in cache:
            map_row = (
                model.query(SP_Customer_SKU_Map)
                .filter(SP_Customer_SKU_Map.CustomerID == customer_id,
                        SP_Customer_SKU_Map.SKU_ID == sku_id)
                .first()
            )
            cache[key_pair] = map_row
        map_row = cache[key_pair]

        if map_row:
            # Mapping exists → must match the provided cust code
            existing_code = (map_row.CustSKUCode or "").strip()
            if existing_code.lower() != cust_code.lower():
                raise ValueError(
                    f"Cust-SKU '{cust_code}' does not match existing mapping for MEC-SKU '{art_val}'. "
                    f"Expected '{existing_code}'."
                )
        else:
            # No (customer, sku) mapping yet → ensure the provided cust code isn't mapped to another SKU
            key_cust = ("cust", customer_id, cust_code)
            if key_cust not in cache:
                mapped_other = (
                    model.query(SP_Customer_SKU_Map.SKU_ID)
                    .filter(SP_Customer_SKU_Map.CustomerID == customer_id,
                            SP_Customer_SKU_Map.CustSKUCode == cust_code)
                    .scalar()
                )
                cache[key_cust] = mapped_other
            mapped_other = cache[key_cust]

            if mapped_other is not None and int(mapped_other) != int(sku_id):
                raise ValueError(
                    f"Cust-SKU '{cust_code}' is already mapped to a different MEC-SKU (SKU_ID={mapped_other})."
                )

            if create_mapping:
                # Create the new mapping
                new_map = SP_Customer_SKU_Map(
                    SKU_ID=sku_id,
                    CustomerID=customer_id,
                    CustSKUCode=cust_code
                )
                model.add(new_map)
                # Update caches so subsequent rows don't query again
                cache[key_pair] = new_map
                cache[key_cust] = sku_id

    # 4) Return what we have (could be None if unresolved and no strict condition to error)
    return sku_id


@bp.route("/choices", methods=["GET"])
def choices():
    customers = [
        {"id": c.CustomerID, "code": c.CustCode, "name": c.CustName, "level": c.LevelType}
        for c in model.query(SP_Customer).filter(SP_Customer.LevelType == 'HO').order_by(SP_Customer.CustName).all()
    ]
    brands = [
        {"name": b[0]} for b in model.query(Brands.BrandName).all() if b[0]
    ]
    return jsonify({"customers": customers, "brands": brands})

def _articles_from_mapping(customer_id: int, brand: str | None):
    """
    Return list of dicts: [{'SKU_ID':..., 'ArticleCode':..., 'CustSKUCode':...}, ...]
    using SP_Customer_SKU_Map (preferred).
    """
    q = (
        model.query(
            SP_Customer_SKU_Map.SKU_ID,
            SP_SKU.ArticleCode,
            SP_Customer_SKU_Map.CustSKUCode
        )
        .join(SP_SKU, SP_SKU.SKU_ID == SP_Customer_SKU_Map.SKU_ID)
        .filter(SP_Customer_SKU_Map.CustomerID == customer_id)
    )
    if brand:
        q = q.filter(SP_SKU.Brand == brand)

    rows = q.all()
    return [
        {
            "SKU_ID": r[0],
            "ArticleCode": r[1],
            "CustSKUCode": r[2],
        }
        for r in rows
        if r[0] is not None and r[1]  # keep only mapped SKUs with article code
    ]

def _articles_from_sellin(customer_id: int, brand: str | None):
    """
    Fallback: mine SP_MCSI_SellIn for distinct Articles sold to this customer,
    then resolve SKU_ID via SP_SKU. We match by customer *name* because your
    SP_MCSI_SellIn.SoldToParty holds names.
    """
    cust = model.query(SP_Customer).get(customer_id)
    if not cust:
        return []

    q = model.query(SP_MCSI_SellIn.Article).filter(SP_MCSI_SellIn.SoldToParty == cust.CustName)
    if brand:
        q = q.filter(SP_MCSI_SellIn.Brand == brand)

    # distinct article codes from sell-in
    articles = {a for (a,) in q.distinct().all() if a}

    if not articles:
        return []

    # resolve to SKU_ID and CustSKUCode (if mapped)
    sku_rows = (
        model.query(SP_SKU.SKU_ID, SP_SKU.ArticleCode)
        .filter(SP_SKU.ArticleCode.in_(list(articles)))
        .all()
    )
    art_to_sku = {ac: sid for (sid, ac) in sku_rows}

    # bring customer mapping where available
    mapped = (
        model.query(SP_Customer_SKU_Map.SKU_ID, SP_Customer_SKU_Map.CustSKUCode)
        .filter(SP_Customer_SKU_Map.CustomerID == customer_id)
        .all()
    )
    sku_to_cust = {sid: ccode for (sid, ccode) in mapped}

    result = []
    for ac in sorted(articles):
        sid = art_to_sku.get(ac)
        if not sid:
            continue
        result.append({
            "SKU_ID": sid,
            "ArticleCode": ac,
            "CustSKUCode": sku_to_cust.get(sid)
        })
    return result

from sqlalchemy import func

def _articles_from_soh(customer_id: int, brand: str | None):
    """
    Pull distinct SKUs from SOH (active rows) for this customer.
    If `brand` is provided, keep ONLY those SKUs whose SP_SKU.Brand matches `brand`.
    Returns [{'SKU_ID', 'ArticleCode', 'CustSKUCode'}].
    """
    brand_norm = (brand or "").strip()
    brand_filter = None
    if brand_norm:
        # case- and space-insensitive comparison on SP_SKU.Brand
        brand_filter = func.lower(func.trim(SP_SKU.Brand)) == func.lower(func.trim(brand_norm))

    # One shot: get distinct SKU_ID + ArticleCode from SOH, optionally filtered by SP_SKU.Brand
    q = (
        model.query(SP_SKU.SKU_ID, SP_SKU.ArticleCode)
        .join(SP_SOH_Detail, SP_SOH_Detail.SKU_ID == SP_SKU.SKU_ID)
        .join(SP_SOH_Uploads, SP_SOH_Uploads.SOHUploadID == SP_SOH_Detail.SOHUploadID)
        .filter(
            SP_SOH_Uploads.CustomerID == customer_id,
            SP_SOH_Detail.IsActive == True
        )
    )
    if brand_filter is not None:
        q = q.filter(brand_filter)

    sku_rows = q.distinct().all()
    if not sku_rows:
        return []

    # Map to dicts we’ll use for CustSKU mapping lookup
    art_by_sku = {sid: ac for (sid, ac) in sku_rows if ac}

    # Bring (customer, sku) mapping for CustSKUCode (if any)
    map_rows = (
        model.query(SP_Customer_SKU_Map.SKU_ID, SP_Customer_SKU_Map.CustSKUCode)
        .filter(
            SP_Customer_SKU_Map.CustomerID == customer_id,
            SP_Customer_SKU_Map.SKU_ID.in_(list(art_by_sku.keys()))
        )
        .all()
    )
    cust_by_sku = {sid: ccode for (sid, ccode) in map_rows}

    # Build output (only SP_SKU-brand-approved SKUs are here)
    out = []
    for sid, ac in art_by_sku.items():
        out.append({
            "SKU_ID": sid,
            "ArticleCode": ac,
            "CustSKUCode": cust_by_sku.get(sid)
        })
    return out

# Dedup/merge utility (prefer mapping, then sell-in, then SOH)
def _merge_items_preferring_mapping(mapping_items, sellin_items, soh_items):
    """
    Return a unified list of dicts (SKU_ID, ArticleCode, CustSKUCode),
    preferring 'mapping_items' values for duplicates by ArticleCode.
    """
    by_article = {}

    # 1) mapping first (authoritative)
    for it in mapping_items or []:
        ac = (it.get("ArticleCode") or "").strip()
        if not ac: 
            continue
        by_article[ac] = {"SKU_ID": it.get("SKU_ID"),
                          "ArticleCode": ac,
                          "CustSKUCode": it.get("CustSKUCode")}

    # 2) then sell-in: add only if not present
    for it in sellin_items or []:
        ac = (it.get("ArticleCode") or "").strip()
        if not ac or ac in by_article:
            continue
        by_article[ac] = {"SKU_ID": it.get("SKU_ID"),
                          "ArticleCode": ac,
                          "CustSKUCode": it.get("CustSKUCode")}

    # 3) then SOH: add only if not present
    for it in soh_items or []:
        ac = (it.get("ArticleCode") or "").strip()
        if not ac or ac in by_article:
            continue
        by_article[ac] = {"SKU_ID": it.get("SKU_ID"),
                          "ArticleCode": ac,
                          "CustSKUCode": it.get("CustSKUCode")}

    # stable order by ArticleCode
    merged = [by_article[k] for k in sorted(by_article.keys())]
    return merged


def _normalize_sellout_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Map user-facing columns to internal ones.
    Accept both new headers and legacy ones.
    """
    rename_map = {}
    cols_lower = {c.lower(): c for c in df.columns}

    # Date
    for k in ["transection date", "transaction date", "documentdate", "date"]:
        if k in cols_lower:
            rename_map[cols_lower[k]] = "DocumentDate"; break

    # Cust SKU
    for k in ["cust-sku", "custsku", "custskucode"]:
        if k in cols_lower:
            rename_map[cols_lower[k]] = "CustSKUCode"; break

    # MEC SKU → we’ll treat as ArticleCode (your internal code)
    for k in ["mec-sku", "articlecode", "article"]:
        if k in cols_lower:
            rename_map[cols_lower[k]] = "ArticleCode"; break

    # Quantities
    if "selloutqty" in cols_lower:
        rename_map[cols_lower["selloutqty"]] = "SellOutQty"
    if "reportedsoh" in cols_lower:
        rename_map[cols_lower["reportedsoh"]] = "ReportedSOH"
    if "rownumber" in cols_lower:
        rename_map[cols_lower["rownumber"]] = "RowNumber"

    df = df.rename(columns=rename_map)
    return df


@bp.route("/template", methods=["GET"])
def template():
    customer_id = request.args.get("customer_id", type=int)
    brand = (request.args.get("brand") or "").strip() or None
    include_soh = request.args.get("include_soh", default="0") in ("1", "true", "True")
    if not customer_id:
        return jsonify(ok=False, error="customer_id is required"), 400

    # Build candidate lists
    items_mapping = _articles_from_mapping(customer_id, brand)           # preferred
    items_sellin  = _articles_from_sellin(customer_id, brand)            # fallback
    items_soh     = _articles_from_soh(customer_id, brand)               # NEW: from SOH

    # Merge → mapping wins; then sell-in; then SOH fills the gaps
    map_rows_full = _merge_items_preferring_mapping(items_mapping, items_sellin, items_soh)

    # (Optional) If you still want the entry grid size to reflect how much we know:
    prefill_rows = max(len(map_rows_full), 50)
    
    cols = ["Transection Date", "Cust-SKU", "MEC-SKU", "SellOutQty"]
    if include_soh:
        cols.append("ReportedSOH")
    
    # decide how many blank entry rows to give the user (purely UX)
    prefill_rows = max(len(map_rows_full), 50)  # or 0 if you want a totally empty sheet
    
    df = pd.DataFrame({
        "Transection Date": [""] * prefill_rows,
        "Cust-SKU":         [""] * prefill_rows,   # <-- leave empty
        "MEC-SKU":          [""] * prefill_rows,   # <-- leave empty (user picks from dropdown)
        "SellOutQty":       [""] * prefill_rows,
        **({"ReportedSOH":  [""] * prefill_rows} if include_soh else {})
    })[cols]
    
    # Date validation window (defaults last year → Today)
    try:
        dr_start = (
            date.fromisoformat(request.args.get("dr_start"))
            if request.args.get("dr_start")
            else date(date.today().year - 1, 1, 1)
        )
        dr_end = (
            date.fromisoformat(request.args.get("dr_end"))
            if request.args.get("dr_end")
            else date.today()
        )
    except Exception:
        dr_start = date(date.today().year - 1, 1, 1)
        dr_end = date.today()
    
    # Build Excel DATE() formulas for validation
    f1 = f"DATE({dr_start.year},{dr_start.month},{dr_start.day})"
    f2 = f"DATE({dr_end.year},{dr_end.month},{dr_end.day})"
    
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as xw:
        # --- Main entry sheet
        df.to_excel(xw, index=False, sheet_name="SellOut")
        wb = xw.book
        ws = wb["SellOut"]
    
        ws.freeze_panes = "A2"
        for ix in range(1, len(cols) + 1):
            ws.column_dimensions[get_column_letter(ix)].width = 18
    
        MAX_ROWS = 1048576  # Excel row limit
    
        # --- Date validation on Column A ---
        dv_date = DataValidation(
            type="date",
            operator="between",
            formula1=f1,
            formula2=f2,
            allow_blank=False,
            showErrorMessage=True,
            errorTitle="Invalid date",
            error=f"Enter a date between {dr_start.isoformat()} and {dr_end.isoformat()}."
        )
        ws.add_data_validation(dv_date)
        dv_date.add(f"A2:A{MAX_ROWS}")
        for r in range(2, min(len(df)+2000, MAX_ROWS) + 1):
            ws[f"A{r}"].number_format = "yyyy-mm-dd"
    
        # --- Non-negative decimal for SellOutQty ---
        qty_col_idx = cols.index("SellOutQty") + 1
        qty_col_letter = get_column_letter(qty_col_idx)
        dv_qty = DataValidation(
            type="decimal",
            operator="greaterThanOrEqual",
            formula1="0",
            allow_blank=False,
            showErrorMessage=True,
            errorTitle="Invalid quantity",
            error="Quantity must be a number ≥ 0."
        )
        ws.add_data_validation(dv_qty)
        dv_qty.add(f"{qty_col_letter}2:{qty_col_letter}{MAX_ROWS}")
    
        # --- Optional non-negative decimal for ReportedSOH ---
        if include_soh:
            soh_col_idx = cols.index("ReportedSOH") + 1
            soh_col_letter = get_column_letter(soh_col_idx)
            dv_soh = DataValidation(
                type="decimal",
                operator="greaterThanOrEqual",
                formula1="0",
                allow_blank=True,
                showErrorMessage=True,
                errorTitle="Invalid SOH",
                error="SOH must be a number ≥ 0 (or leave blank)."
            )
            ws.add_data_validation(dv_soh)
            dv_soh.add(f"{soh_col_letter}2:{soh_col_letter}{MAX_ROWS}")
    
        # --- SKU MAP sheet (MEC-SKU, Cust-SKU) ---
        # If no explicit mapping rows exist, still create the sheet with headers.
        map_df = pd.DataFrame(map_rows_full) if map_rows_full else pd.DataFrame(columns=["SKU_ID","ArticleCode","CustSKUCode"])
        # Keep just the two codes for the user
        sku_map_view = pd.DataFrame({
            "MEC-SKU": map_df.get("ArticleCode", pd.Series(dtype=str)),
            "Cust-SKU": map_df.get("CustSKUCode", pd.Series(dtype=str)),
        })
        sku_map_view.to_excel(xw, index=False, sheet_name="SKU_Map")
        wmap = wb["SKU_Map"]
        wmap.freeze_panes = "A2"
        wmap.column_dimensions["A"].width = 26
        wmap.column_dimensions["B"].width = 26
    
        # Determine ranges for list validations (limit to actual rows; if none, point to empty row 2)
        map_rows_count = max(2, len(sku_map_view) + 1)  # header is row 1; data starts row 2
        mec_list_range  = f"'SKU_Map'!$A$2:$A${map_rows_count}"
        cust_list_range = f"'SKU_Map'!$B$2:$B${map_rows_count}"
    
        # Column letters on SellOut sheet
        cust_col_letter = get_column_letter(cols.index("Cust-SKU") + 1)  # usually "B"
        mec_col_letter  = get_column_letter(cols.index("MEC-SKU")  + 1)  # usually "C"

        # --- Auto-populate Cust-SKU from MEC-SKU via VLOOKUP (user can overwrite) ---
        # Use the actual list range instead of whole column for speed
        lookup_range = f"'SKU_Map'!$A$2:$B${map_rows_count}"
        
        # How many entry rows to pre-provision with the formula
        AUTO_ROWS = max(1000, len(df) + 200)  # tweak as you like
        
        for r in range(2, min(AUTO_ROWS + 1, MAX_ROWS + 1)):
            # If MEC-SKU is blank -> blank; else try VLOOKUP to get Cust-SKU
            ws[f"{cust_col_letter}{r}"].value = (
                f'=IFERROR(IF({mec_col_letter}{r}="","",'
                f'VLOOKUP({mec_col_letter}{r},{lookup_range},2,FALSE)),"")'
            )
        
        # ------------- Validations & visual rules -------------
        # 1) MEC-SKU must be from MEC list (dropdown restriction)
        dv_mec_list = DataValidation(
            type="list",
            formula1=mec_list_range,
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Unknown MEC-SKU",
            error="Choose a valid MEC-SKU from the dropdown (SKU_Map sheet)."
        )
        ws.add_data_validation(dv_mec_list)
        dv_mec_list.add(f"{mec_col_letter}2:{mec_col_letter}{MAX_ROWS}")
    
        # 2) If Customer SKU is NOT in the mapping list, then MEC-SKU must be provided AND be valid.
        # Custom rule attached to the Cust-SKU column:
        # =IF( ISNUMBER(MATCH(B2, SKU_Map!$B:$B, 0)),
        #      TRUE,
        #      AND( C2<>"", ISNUMBER(MATCH(C2, SKU_Map!$A:$A, 0)) )
        #    )
        dv_dep = DataValidation(
            type="custom",
            formula1=(
                f"IF(ISNUMBER(MATCH({cust_col_letter}2,SKU_Map!$B:$B,0)),"
                f"TRUE,"
                f"AND({mec_col_letter}2<>\"\",ISNUMBER(MATCH({mec_col_letter}2,SKU_Map!$A:$A,0))))"
            ),
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Missing required MEC-SKU",
            error="This Customer SKU is not in the mapping. Please enter a valid MEC-SKU in the next column."
        )
        ws.add_data_validation(dv_dep)
        dv_dep.add(f"{cust_col_letter}2:{cust_col_letter}{MAX_ROWS}")
    
        # 3) Red highlight when a typed Customer SKU doesn’t exist in mapping
        red_fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
        red_rule = FormulaRule(
            formula=[f'AND({cust_col_letter}2<>"",ISNA(MATCH({cust_col_letter}2,SKU_Map!$B:$B,0)))'],
            stopIfTrue=True,
            fill=red_fill
        )
        ws.conditional_formatting.add(f"{cust_col_letter}2:{cust_col_letter}{MAX_ROWS}", red_rule)
    
        # 4) Red highlight when a typed MEC-SKU isn’t in the MEC list
        red_rule_mec = FormulaRule(
            formula=[f'AND({mec_col_letter}2<>"",ISNA(MATCH({mec_col_letter}2,SKU_Map!$A:$A,0)))'],
            stopIfTrue=True,
            fill=red_fill
        )
        ws.conditional_formatting.add(f"{mec_col_letter}2:{mec_col_letter}{MAX_ROWS}", red_rule_mec)
    
        # Info sheet
        meta = pd.DataFrame(
            {
                "Key": ["GeneratedAt", "CustomerID", "BrandFilter", "Source", "DateWindow"],
                "Value": [
                    datetime.now(TZ_RIYADH).isoformat(timespec="seconds") + "Z",
                    customer_id,
                    brand or "(all)",
                    "Map" if items_mapping else "Blank",
                    f"{dr_start}..{dr_end}",
                ],
            }
        )
        meta.to_excel(xw, index=False, sheet_name="Info")
    
    bio.seek(0)
    return send_file(
        bio,
        as_attachment=True,
        download_name=f"sellout_template_{customer_id}{('_' + brand) if brand else ''}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@bp.route("/upload", methods=["POST"])
def upload():
    try:
        # -------- 0) Inputs & file presence --------
        customer_id = request.form.get("customer_id", type=int)
        level_type  = (request.form.get("level_type")  or "").strip()
        upload_type = (request.form.get("upload_type") or "").strip()
        brand       = (request.form.get("brand")       or "").strip() or None
        created_by  = (request.form.get("created_by")  or "").strip() or None
        fs = request.files.get("file")

        if not customer_id or not level_type or not upload_type:
            return jsonify(ok=False, error="customer_id, level_type, upload_type are required"), 400
        if level_type not in ("HO","Branch"):
            return jsonify(ok=False, error="level_type must be 'HO' or 'Branch'"), 400
        if upload_type not in ("Customer-Format","Company-Format"):
            return jsonify(ok=False, error="upload_type must be 'Customer-Format' or 'Company-Format'"), 400
        if not fs or not _allowed(fs.filename):
            return jsonify(ok=False, error="Upload a .xlsx/.xls/.csv file"), 400

        # -------- 1) Duplicate file (soft) --------
        file_hash = _sha256_fs(fs)
        dup = (model.query(SP_SellOutUploads.UploadID)
               .filter(SP_SellOutUploads.SourceFileHash == file_hash)
               # If you want per-customer scope, also filter CustomerID here.
               .first())
        if dup:
            return jsonify(ok=False, error="This exact file was already uploaded."), 400

        # -------- 2) Parse & validate file (in memory only) --------
        df = _load_df(fs).copy()
        df.columns = [c.strip() for c in df.columns]
        if "DocumentDate" not in df.columns:
            return jsonify(ok=False, error="File must contain 'DocumentDate' column."), 400

        df["DocumentDate"] = pd.to_datetime(df["DocumentDate"], errors="coerce").dt.date
        if "SellOutQty" in df.columns:   df["SellOutQty"] = pd.to_numeric(df["SellOutQty"], errors="coerce")
        if "ReportedSOH" in df.columns:  df["ReportedSOH"] = pd.to_numeric(df["ReportedSOH"], errors="coerce")
        if "RowNumber"  not in df.columns: df["RowNumber"] = range(1, len(df)+1)

        df = df.dropna(subset=["DocumentDate"])
        if df.empty:
            return jsonify(ok=False, error="No valid rows (DocumentDate missing after parsing)"), 400

        # ---- Sell-Out hard validations: future dates, negative/decimal qty ----
        today_riyadh = datetime.now(TZ_RIYADH).date() if TZ_RIYADH else date.today()
        errors = []
        
        # 1) Future date check
        future_rows = df.loc[df["DocumentDate"] > today_riyadh, ["RowNumber", "DocumentDate"]]
        if not future_rows.empty:
            msgs = [f"Row {int(r.RowNumber)}: DocumentDate {r.DocumentDate} cannot be in the future."
                    for _, r in future_rows.iterrows()]
            errors.extend(msgs)
        
        # 2) Negative qty check (only validate where qty is provided)
        if "SellOutQty" in df.columns:
            neg_rows = df.loc[df["SellOutQty"].notna() & (df["SellOutQty"] < 0), ["RowNumber","SellOutQty"]]
            if not neg_rows.empty:
                msgs = [f"Row {int(r.RowNumber)}: SellOutQty cannot be negative (got {r.SellOutQty})."
                        for _, r in neg_rows.iterrows()]
                errors.extend(msgs)
        
            # 3) Decimal qty check (qty must be a whole number)
            dec_mask = df["SellOutQty"].notna() & ((df["SellOutQty"] % 1) != 0)
            dec_rows = df.loc[dec_mask, ["RowNumber","SellOutQty"]]
            if not dec_rows.empty:
                msgs = [f"Row {int(r.RowNumber)}: SellOutQty must be a whole number (got {r.SellOutQty})."
                        for _, r in dec_rows.iterrows()]
                errors.extend(msgs)
        
        if errors:
            return jsonify(ok=False, error="; ".join(errors)), 400
        
        
        period_start = df["DocumentDate"].min()
        period_end   = df["DocumentDate"].max()

        # Resolve SKUs & build records
        cache = {}
        records = []
        for idx, r in df.iterrows():
            try:
                sku_id = _resolve_sku_id(r, customer_id, cache)
            except ValueError as ve:
                row_no = int(r.get("RowNumber", idx + 2))
                return jsonify(ok=False, error=f"Row {row_no}: {ve}"), 400
            if sku_id is None:
                continue
            qty = r.get("SellOutQty")
            if pd.isna(qty):
                continue
            records.append({
                "DocumentDate": r["DocumentDate"],
                "SKU_ID": int(sku_id),
                "CustSKUCode": (str(r["CustSKUCode"]).strip()
                                if "CustSKUCode" in r and pd.notna(r["CustSKUCode"]) else None),
                "SellOutQty": float(qty),
                "ReportedSOH": (float(r["ReportedSOH"])
                                if "ReportedSOH" in r and pd.notna(r["ReportedSOH"]) else None),
                "RowNumber": int(r["RowNumber"])
            })
        if not records:
            return jsonify(ok=False, error="No valid rows after SKU/Qty resolution"), 400


        # ---- SELL-IN / SOH HISTORY CHECK ----
        # We allow upload if each SKU in this file has:
        #   EITHER (a) prior Sell-In for this customer
        #   OR     (b) an active/in-Active SOH entry for this customer.

        # 1) Collect distinct SKUs from this file
        sku_ids = {rec["SKU_ID"] for rec in records}

        cust_name = model.query(SP_Customer.CustName).filter(SP_Customer.CustomerID == customer_id).scalar()
        if not cust_name:
            return jsonify(ok=False, error=f"Unknown customer_id={customer_id}"), 400

        # Map SKU → ArticleCode
        sku_rows = (
            model.query(SP_SKU.SKU_ID, SP_SKU.ArticleCode)
                .filter(SP_SKU.SKU_ID.in_(sku_ids))
                .all()
        )
        sku_to_article = {sid: ac for sid, ac in sku_rows}
        articles_needed = {ac for ac in sku_to_article.values() if ac}

        # (a) Find which have Sell-In (optionally brand-filtered)
        q_si = model.query(SP_MCSI_SellIn.Article).filter(SP_MCSI_SellIn.SoldToParty == cust_name)
        if brand:
            q_si = q_si.filter(SP_MCSI_SellIn.Brand == brand)

        found_articles = {
            a for (a,) in q_si.filter(SP_MCSI_SellIn.Article.in_(list(articles_needed))).distinct().all()
        }

        # SKUs that do NOT have Sell-In
        no_sellin_skus = {sid for sid, ac in sku_to_article.items() if ac not in found_articles}

        if no_sellin_skus:
            # (b) Allow if the SKU appears in SOH for this customer.
            # NOTE: We scope SOH check to only SKUs from this file for efficiency.
            soh_sku_ids = {
                sid for (sid,) in (
                    model.query(SP_SOH_Detail.SKU_ID)
                        .join(SP_SOH_Uploads, SP_SOH_Uploads.SOHUploadID == SP_SOH_Detail.SOHUploadID)
                        .filter(
                            SP_SOH_Uploads.CustomerID == customer_id,
                            SP_SOH_Detail.SKU_ID.in_(list(no_sellin_skus))
                        )
                        .distinct()
                        .all()
                )
            }

            # Only block SKUs that have neither Sell-In nor SOH
            still_blocked = no_sellin_skus - soh_sku_ids

            if still_blocked:
                # Collect offending rows
                bad_rows = sorted({rec["RowNumber"] for rec in records if rec["SKU_ID"] in still_blocked})
                sku_info = ", ".join(f"{sku_to_article[sid]}(SKU_ID={sid})" for sid in still_blocked)

                return jsonify(
                    ok=False,
                    error=(
                        "Sell-In / SOH not found for "
                        f"{len(still_blocked)} SKU(s) [{sku_info}]. "
                        f"Check Excel rows (excluding header) ROWS: {', '.join(map(str, bad_rows))}."
                    )
                ), 400
        
        
        # Precompute overlap candidates (read-only)
        hdr_ids_q = (model.query(SP_SellOutUploads.UploadID)
                     .filter(SP_SellOutUploads.CustomerID==customer_id,
                             SP_SellOutUploads.LevelType==level_type,
                             SP_SellOutUploads.UploadType==upload_type))
        if brand is not None:
            hdr_ids_q = hdr_ids_q.filter(SP_SellOutUploads.Brand==brand)
        hdr_ids = [h[0] for h in hdr_ids_q.all()]

        prior_headers = []
        if hdr_ids:
            prior_headers = (model.query(SP_SellOutUploads)
                             .filter(SP_SellOutUploads.UploadID.in_(hdr_ids),
                                     SP_SellOutUploads.PeriodStart <= period_end,
                                     SP_SellOutUploads.PeriodEnd   >= period_start)
                             .all())

        # -------- 3) ATOMIC WRITE SECTION --------
        # Everything inside runs in one DB transaction. Any exception → rollback all.
        with _begin_tx(model): # SQLAlchemy session transaction context
            # Header (unique hash may raise IntegrityError if duplicate under race)
            header = SP_SellOutUploads(
                CustomerID=customer_id,
                LevelType=level_type,
                UploadType=upload_type,
                Brand=brand,
                DocumentDate=period_end, #lagacy support
                PeriodStart=period_start,
                PeriodEnd=period_end,
                Status="Draft",
                CreatedBy=created_by,
                CreatedAt=datetime.utcnow(),
                SourceFileName=fs.filename,
                SourceFileHash=file_hash
            )
            model.add(header)
            model.flush()  # ensure header.UploadID available

            # Link overlapping prior headers without downgrading Posted ones
            superseded_ids = []
            for h in prior_headers:
                # Always link who superseded who (for traceability)
                h.SupersededByUploadID = header.UploadID
            
                # Only change status for non-posted drafts/rejected (optional: use a neutral 'Superseded')
                if h.Status in ("Draft", "Rejected"):
                    # If you already have a 'Superseded' status in your enum / UI, use it:
                    h.Status = "Superseded"
                    # Otherwise, you can mark as 'Rejected' to make them non-actionable:
                    # h.Status = "Rejected"
            
                superseded_ids.append(h.UploadID)

            # Deactivate prior details in the window (soft replace)
            deactivated = 0
            if hdr_ids:
                deactivated = (model.query(SP_MCSI_SellOut)
                               .filter(SP_MCSI_SellOut.UploadID.in_(hdr_ids),
                                       SP_MCSI_SellOut.DocumentDate >= period_start,
                                       SP_MCSI_SellOut.DocumentDate <= period_end)
                               .update({SP_MCSI_SellOut.IsActive: False},
                                       synchronize_session=False))

            # Insert new details
            detail_objs = [SP_MCSI_SellOut(
                UploadID=header.UploadID,
                DocumentDate=r["DocumentDate"],
                SKU_ID=r["SKU_ID"],
                CustSKUCode=r["CustSKUCode"],
                SellOutQty=r["SellOutQty"],
                ReportedSOH=r["ReportedSOH"],
                RowNumber=r["RowNumber"],
                IsActive=True
            ) for r in records]
            model.add_all(detail_objs)

            # Audit
            model.add(SP_SellOutUploadAudit(
                Action="UPLOAD" if not superseded_ids else "REPLACE",
                CustomerID=customer_id, LevelType=level_type,
                UploadType=upload_type, Brand=brand,
                PeriodStart=period_start, PeriodEnd=period_end,
                NewUploadID=header.UploadID,
                SupersededUploadIDs=",".join(map(str, superseded_ids)) if superseded_ids else None,
                DeactivatedRows=int(deactivated),
                InsertedRows=len(detail_objs),
                SourceFileName=fs.filename,
                SourceFileHash=file_hash,
                PerformedBy=created_by
            ))

        # ---- If we got here, everything committed atomically ----
        return jsonify(
            ok=True,
            upload_id=header.UploadID,
            period=[str(period_start), str(period_end)],
            inserted=len(records),
            deactivated=int(deactivated),
            superseded=superseded_ids
        )

    except IntegrityError:
        # Duplicate SHA or other constraint violation → nothing persisted thanks to transaction
        model.rollback()
        return jsonify(ok=False, error="Duplicate file detected (same SHA-256)."), 400
    except Exception as e:
        model.rollback()
        current_app.logger.exception("Sell-Out upload failed")
        return jsonify(ok=False, error=str(e)), 500

@bp.route("/", methods=["GET", "POST"])
def front():
    return render_template('./sell_out/sell_out.html')